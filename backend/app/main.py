from fastapi import FastAPI, HTTPException, Depends, status, UploadFile, File, Form, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.security import OAuth2PasswordBearer
from typing import List, Dict, Any, Optional
from uuid import uuid4
from datetime import datetime, timezone, timedelta
from pathlib import Path
from shutil import copyfile, move, which
from zipfile import ZipFile, BadZipFile, ZIP_DEFLATED
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.reader import workbook as _wb_reader
from sqlalchemy import text as _sa_text
from sqlalchemy.orm import Session
from sqlalchemy.orm.attributes import flag_modified
import bcrypt
from jose import JWTError, jwt
import io
import subprocess
import os
import tempfile
import json
import base64
import re
import sys
import traceback
import unicodedata
import anthropic
from docxtpl import DocxTemplate

from dotenv import load_dotenv
load_dotenv()

from pydantic import BaseModel as _PydanticBase
from .database import get_db
from . import models, schemas
from .amureba_mapper import AmurebaMappingService


class _LcaMaterialEditPayload(_PydanticBase):
    """Payload étendu pour PATCH /lca/materials/{id} (nom et catégorie inclus)."""
    name: Optional[str] = None
    category: Optional[str] = None
    prix: Optional[float] = None
    valeur_r: Optional[float] = None
    flux_reference: Optional[float] = None


# Patch openpyxl: ignore corrupted pivot caches in AMUREBA template
_orig_pivot_caches = _wb_reader.WorkbookParser.pivot_caches.fget


def _safe_pivot_caches(self):
    try:
        return _orig_pivot_caches(self)
    except Exception:
        return {}


_wb_reader.WorkbookParser.pivot_caches = property(_safe_pivot_caches)


# ==============================
# CONFIG
# ==============================
_SOFFICE_MACOS = "/Applications/LibreOffice.app/Contents/MacOS/soffice"
_SOFFICE_WIN  = r"C:\Program Files\LibreOffice\program\soffice.exe"
SOFFICE = (
    which("libreoffice")
    or which("soffice")
    or (_SOFFICE_WIN  if Path(_SOFFICE_WIN).exists()  else None)
    or (_SOFFICE_MACOS if Path(_SOFFICE_MACOS).exists() else None)
)

BASE_DIR = Path(__file__).parent

TEMPLATE_FILE = BASE_DIR / "templates" / "audit_template.xlsx"
EXCEL_DIR = BASE_DIR / "excel"
EXCEL_DIR.mkdir(exist_ok=True)

print(f"[DEBUG] BASE_DIR: {BASE_DIR}", flush=True)
print(f"[DEBUG] TEMPLATE_FILE: {TEMPLATE_FILE}", flush=True)
print(f"[DEBUG] Template exists: {TEMPLATE_FILE.exists()}", flush=True)
print(f"[DEBUG] Templates dir contents: {list((BASE_DIR / 'templates').iterdir()) if (BASE_DIR / 'templates').exists() else 'DIR NOT FOUND'}", flush=True)

REPORT_TEMPLATE_FILE = BASE_DIR / "templates" / "report_template.docx"
REPORT_DIR = BASE_DIR / "reports"
REPORT_DIR.mkdir(exist_ok=True)

SHEET_NAME = "2023"

TITLE_ROWS = {
    "operational": 5,
    "buildings": 8,
    "transport": 11,
    "utility": 14,
}

SECTION_START_ROW = {
    "operational": 6,
    "buildings": 9,
    "transport": 12,
    "utility": 15,
}

MAX_ROWS_PER_SECTION = 2
INFLUENCE_START_ROW = 6
INFLUENCE_MAX_ROWS = 8

INDICES_CELLS = {
    "primary": {"IEE": "B43", "IC": "B44", "iSER": "B45"},
    "secondary": {"AEE": "B49", "iCO2": "B50", "ACO2": "B51"},
}


def _safe_filename(name: str) -> str:
    name = unicodedata.normalize('NFKD', name).encode('ascii', 'ignore').decode('ascii')
    name = re.sub(r'\s+', '_', name)
    name = re.sub(r'[^\w\-]', '', name)
    return name or 'projet'


# ──────────────────────────────────────────────────────────────────────────────
# ZIPFILE-BASED XLSX WRITER (fixes named-range corruption from openpyxl)
# ──────────────────────────────────────────────────────────────────────────────

_NS_MAIN  = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NS_R     = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_NS_MC    = "http://schemas.openxmlformats.org/markup-compatibility/2006"
_NS_RELS  = "http://schemas.openxmlformats.org/package/2006/relationships"
_NS_X14AC = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac"
_NS_XR    = "http://schemas.microsoft.com/office/spreadsheetml/2014/revision"
_NS_XR2   = "http://schemas.microsoft.com/office/spreadsheetml/2015/revision2"
_NS_XR3   = "http://schemas.microsoft.com/office/spreadsheetml/2016/revision3"

for _pfx, _uri in [
    ("",      _NS_MAIN),
    ("r",     _NS_R),
    ("mc",    _NS_MC),
    ("x14ac", _NS_X14AC),
    ("xr",    _NS_XR),
    ("xr2",   _NS_XR2),
    ("xr3",   _NS_XR3),
]:
    ET.register_namespace(_pfx, _uri)


def _build_sheet_path_map(zf: ZipFile) -> Dict[str, str]:
    """Return {sheet_name: 'xl/worksheets/sheetN.xml'} from workbook.xml + rels."""
    wb_root = ET.fromstring(zf.read("xl/workbook.xml"))
    name_to_rid: Dict[str, str] = {}
    for el in wb_root.iter(f"{{{_NS_MAIN}}}sheet"):
        name = el.get("name")
        rid  = el.get(f"{{{_NS_R}}}id")
        if name and rid:
            name_to_rid[name] = rid

    rels_root = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    rid_to_path: Dict[str, str] = {}
    for rel in rels_root.iter(f"{{{_NS_RELS}}}Relationship"):
        rid    = rel.get("Id")
        target = rel.get("Target", "")
        if rid:
            path = target if target.startswith("xl/") else f"xl/{target}"
            rid_to_path[rid] = path

    return {n: rid_to_path[r] for n, r in name_to_rid.items() if r in rid_to_path}


def _col_num(col: str) -> int:
    """'A'→1, 'B'→2, 'AA'→27"""
    n = 0
    for c in col.upper():
        n = n * 26 + (ord(c) - 64)
    return n


def _cell_col_num(ref: str) -> int:
    m = re.match(r'([A-Z]+)', str(ref).upper())
    return _col_num(m.group(1)) if m else 0


def _patch_sheet_xml(xml_bytes: bytes, changes: Dict[str, Any]) -> bytes:
    """
    Write cell values into a worksheet XML using ElementTree.
    - Numbers  → plain <v>N</v>, t attribute removed (default numeric type).
    - Strings  → t="inlineStr" + <is><t>text</t></is>  (no sharedStrings.xml change).
    - Formulas on targeted cells are removed so our constants are not overwritten.
    - Style (s=) attribute on existing cells is preserved.
    - ET auto-escapes &, <, > in text content.
    """
    root = ET.fromstring(xml_bytes)
    NS = _NS_MAIN

    sheet_data = root.find(f"{{{NS}}}sheetData")
    if sheet_data is None:
        return xml_bytes

    # Build row_num → element map
    row_map: Dict[int, ET.Element] = {}
    for row_el in sheet_data:
        try:
            row_map[int(row_el.get("r", 0))] = row_el
        except (ValueError, TypeError):
            pass

    for cell_ref, value in changes.items():
        m = re.match(r'([A-Z]+)(\d+)', str(cell_ref).upper())
        if not m:
            continue
        row_num = int(m.group(2))

        # Find or create row
        if row_num not in row_map:
            row_el = ET.SubElement(sheet_data, f"{{{NS}}}row")
            row_el.set("r", str(row_num))
            row_map[row_num] = row_el
        else:
            row_el = row_map[row_num]

        # Find or create cell
        cell_el: Optional[ET.Element] = None
        for c in row_el:
            if c.get("r") == cell_ref:
                cell_el = c
                break
        if cell_el is None:
            cell_el = ET.SubElement(row_el, f"{{{NS}}}c")
            cell_el.set("r", cell_ref)

        # Remove ALL existing children (formulas + cached values)
        for child in list(cell_el):
            cell_el.remove(child)

        if isinstance(value, (int, float)) and not isinstance(value, bool):
            # Numeric constant — remove t so Excel treats cell as number
            cell_el.attrib.pop("t", None)
            ET.SubElement(cell_el, f"{{{NS}}}v").text = str(value)
        elif value is not None and str(value).strip():
            # String — inline string (no sharedStrings.xml involvement)
            cell_el.set("t", "inlineStr")
            is_el = ET.SubElement(cell_el, f"{{{NS}}}is")
            ET.SubElement(is_el, f"{{{NS}}}t").text = str(value)
        else:
            # Empty cell — clear type, leave no value child
            cell_el.attrib.pop("t", None)

    # Re-sort rows by r, and cells within each row by column index
    sorted_rows = sorted(list(sheet_data), key=lambda e: int(e.get("r") or 0))
    for child in list(sheet_data):
        sheet_data.remove(child)
    for row_el in sorted_rows:
        cells = sorted(list(row_el), key=lambda e: _cell_col_num(e.get("r") or "A0"))
        for child in list(row_el):
            row_el.remove(child)
        for cell in cells:
            row_el.append(cell)
        sheet_data.append(row_el)

    body = ET.tostring(root, encoding="unicode")
    return b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n' + body.encode("utf-8")


def _apply_changes_to_template(
    template_path: Path,
    sheet_changes: Dict[str, Dict[str, Any]],
) -> bytes:
    """
    Low-level: copy template zip, patch only the listed worksheet cells, return bytes.
    workbook.xml and sharedStrings.xml are copied byte-for-byte (no corruption).
    """
    with ZipFile(template_path, "r") as zf_in:
        sheet_path_map = _build_sheet_path_map(zf_in)
        path_to_sheet  = {v: k for k, v in sheet_path_map.items()}

        out_buf = io.BytesIO()
        with ZipFile(out_buf, "w", compression=ZIP_DEFLATED) as zf_out:
            for item in zf_in.infolist():
                data = zf_in.read(item.filename)
                sheet_name = path_to_sheet.get(item.filename)
                if sheet_name and sheet_name in sheet_changes:
                    try:
                        data = _patch_sheet_xml(data, sheet_changes[sheet_name])
                    except Exception as exc:
                        print(f"[WARN] _patch_sheet_xml({item.filename}): {exc}", flush=True)
                zf_out.writestr(item, data)

    return out_buf.getvalue()


def _write_template_zipfile(
    template_path: Path,
    entity_name: str,
    actions: List[Dict[str, Any]],
) -> bytes:
    """
    High-level: build sheet_changes from entity_name + Claude actions list,
    then delegate to _apply_changes_to_template.
    Both text (inlineStr) and numeric values are injected.
    workbook.xml / sharedStrings.xml remain byte-identical to template.
    """
    sheet_changes = _build_prefill_sheet_changes(entity_name, actions)
    return _apply_changes_to_template(template_path, sheet_changes)


def _build_prefill_sheet_changes(
    entity_name: str,
    actions: List[Dict[str, Any]],
) -> Dict[str, Dict[str, Any]]:
    """
    Build the exact worksheet patch payload used for AMUREBA prefill.
    Reused by both the full IA prefill and the checkbox-based apply flow so
    both routes generate files in the same safe way.
    """
    sheet_changes: Dict[str, Dict[str, Any]] = {}

    def _s(sheet: str, cell: str, val: Any) -> None:
        sheet_changes.setdefault(sheet, {})[cell] = val

    if entity_name:
        _s("Paramètres", "B3", entity_name)
        _s("Entête", "C15", entity_name)

    for i, action in enumerate(actions[:9], start=1):
        sh = action.get("sheet") or f"AA{i}"

        # Keep the common header injection that already worked with the
        # original prefill flow.
        if entity_name:
            _s(sh, "B1", entity_name)

        # Reapply the old safe defaults that produced stable AMUREBA files,
        # then let selected user values override them.
        _s(sh, "B9", action.get("intitule") or "")
        _s(sh, "B13", action.get("type_amelioration") or "")
        _s(sh, "F27", action.get("classification") or "B")
        _s(sh, "K22", action.get("ets") or "NON")
        _s(sh, "K23", action.get("deduction_fiscale") or "OUI")

        optional_text_fields = [
            ("B16", "situation_existante"),
            ("B20", "description"),
        ]
        for cell, key in optional_text_fields:
            value = action.get(key)
            if value is None:
                continue
            _s(sh, cell, str(value))

        for cell, key in [
            ("G61", "investissement_k_eur"),
            ("G77", "economie_energie_mwh_an"),
            ("G87", "economie_co2_kg_an"),
            ("K18", "duree_amortissement"),
        ]:
            if key not in action:
                continue
            val = action.get(key)
            if val is None:
                continue
            try:
                _s(sh, cell, float(val))
            except (TypeError, ValueError):
                pass

    return sheet_changes


# ==============================
# AUTH CONFIG
# ==============================
SECRET_KEY = os.environ.get("SECRET_KEY", "change-me-in-production")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_DAYS = 7

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="/auth/login")


def _hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()


def _verify_password(plain: str, hashed: str) -> bool:
    return bcrypt.checkpw(plain.encode(), hashed.encode())


def _create_access_token(user_id: str) -> str:
    expire = datetime.now(timezone.utc) + timedelta(days=ACCESS_TOKEN_EXPIRE_DAYS)
    return jwt.encode({"sub": user_id, "exp": expire}, SECRET_KEY, algorithm=ALGORITHM)


def get_current_user(
    token: str = Depends(oauth2_scheme), db: Session = Depends(get_db)
) -> models.User:
    exc = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Token invalide ou expiré",
        headers={"WWW-Authenticate": "Bearer"},
    )
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        if not user_id:
            raise exc
    except JWTError:
        raise exc
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise exc
    return user


# ==============================
# FASTAPI APP
# ==============================
app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=[
        "http://localhost:5173",
        "http://localhost:5174",
        "http://127.0.0.1:5173",
        "http://127.0.0.1:5174",
    ],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
def _ensure_extra_project_columns():
    """Ajoute les colonnes optionnelles sur projects si elles n'existent pas encore."""
    from sqlalchemy import text
    from .database import engine
    stmts = [
        "ALTER TABLE projects ADD COLUMN IF NOT EXISTS excel_summary JSONB",
        "ALTER TABLE projects ADD COLUMN IF NOT EXISTS prefill_summary JSONB",
        "ALTER TABLE projects ADD COLUMN IF NOT EXISTS prefilled_excel BYTEA",
        "ALTER TABLE projects ADD COLUMN IF NOT EXISTS prefilled_at TEXT",
        # Nouvelle table historique (idempotent)
        """CREATE TABLE IF NOT EXISTS plan_amelioration_history (
            id TEXT PRIMARY KEY,
            project_id TEXT NOT NULL REFERENCES projects(id) ON DELETE CASCADE,
            owner_id TEXT NOT NULL REFERENCES users(id) ON DELETE CASCADE,
            action_type TEXT NOT NULL,
            changes JSONB,
            created_at TEXT NOT NULL
        )""",
    ]
    with engine.begin() as conn:
        for stmt in stmts:
            conn.execute(text(stmt))


# ==============================
# ROUTES: AUTH
# ==============================
@app.post("/auth/register", status_code=201)
def register(payload: schemas.UserCreate, db: Session = Depends(get_db)):
    if db.query(models.User).filter(models.User.email == payload.email).first():
        raise HTTPException(status_code=400, detail="Email déjà utilisé")
    user = models.User(
        id=str(uuid4()),
        full_name=payload.full_name,
        email=payload.email,
        hashed_password=_hash_password(payload.password),
    )
    db.add(user)
    db.commit()
    return {"message": "Compte créé avec succès"}


@app.post("/auth/login")
def login(payload: schemas.UserLogin, db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.email == payload.email).first()
    if not user or not _verify_password(payload.password, user.hashed_password):
        raise HTTPException(status_code=401, detail="Email ou mot de passe incorrect")
    token = _create_access_token(user.id)
    return {
        "access_token": token,
        "token_type": "bearer",
        "user": {"id": user.id, "email": user.email, "full_name": user.full_name},
    }


# ==============================
# HELPERS EXCEL
# ==============================
def _to_number(v: Any):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return v
    s = str(v).strip()
    if s == "":
        return None
    s = s.replace(" ", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _set_cell(ws, addr: str, value: Any, numeric: bool = False):
    if numeric:
        n = _to_number(value)
        ws[addr].value = n
        if n is not None:
            ws[addr].number_format = "0.00"
    else:
        ws[addr].value = value if value is not None else ""


def _get_sheet(wb):
    if SHEET_NAME in wb.sheetnames:
        return wb[SHEET_NAME]
    print(f"[WARN] Feuille '{SHEET_NAME}' introuvable. Feuilles: {wb.sheetnames}. Utilisation: '{wb.active.title}'.")
    return wb.active


def is_valid_excel(path: Path) -> bool:
    try:
        ZipFile(path).close()
        return True
    except Exception:
        return False


def recalc_excel_in_place(excel_path: Path) -> None:
    if not excel_path.exists() or not is_valid_excel(excel_path):
        return
    if not SOFFICE:
        return

    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir = Path(tmpdir)
        _kwargs = {}
        if sys.platform == "win32":
            _kwargs["creationflags"] = subprocess.CREATE_NO_WINDOW
        subprocess.run(
            [SOFFICE, "--headless", "--nologo", "--nolockcheck", "--norestore",
             "--convert-to", "xlsx", "--outdir", str(tmpdir), str(excel_path)],
            check=True,
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            **_kwargs,
        )
        generated = tmpdir / excel_path.name
        if not generated.exists():
            cands = list(tmpdir.glob("*.xlsx"))
            if cands:
                generated = cands[0]
        if generated.exists() and is_valid_excel(generated):
            copyfile(str(generated), str(excel_path))
            generated.unlink()


def _ensure_excel(project, db: Session) -> None:
    """Recrée le fichier Excel depuis le template + données audit si absent (filesystem éphémère sur Render)."""
    excel_path = EXCEL_DIR / project.excel_file
    print(f"[DEBUG] _ensure_excel: excel_path={excel_path}", flush=True)
    print(f"[DEBUG] _ensure_excel: excel_path.exists()={excel_path.exists()}", flush=True)
    print(f"[DEBUG] _ensure_excel: EXCEL_DIR exists={EXCEL_DIR.exists()}, writable={os.access(EXCEL_DIR, os.W_OK)}", flush=True)
    if excel_path.exists() and is_valid_excel(excel_path):
        return
    if excel_path.exists():
        print(f"[DEBUG] _ensure_excel: fichier corrompu, suppression...", flush=True)
        excel_path.unlink()
    print(f"[DEBUG] _ensure_excel: fichier absent ou corrompu, copie du template...", flush=True)
    try:
        copyfile(TEMPLATE_FILE, excel_path)
        print(f"[DEBUG] _ensure_excel: copie réussie, fichier créé={excel_path.exists()}", flush=True)
    except Exception as e:
        print(f"[DEBUG] _ensure_excel: ERREUR copyfile: {e}", flush=True)
        raise
    audit = db.query(models.Audit).filter(models.Audit.project_id == project.id).first()
    if audit:
        write_audit_to_excel(project, _audit_to_data(audit))


def write_audit_to_excel(project, audit_data: Dict[str, Any]) -> None:
    excel_path = EXCEL_DIR / project.excel_file
    copyfile(TEMPLATE_FILE, excel_path)

    wb = load_workbook(excel_path, keep_links=False)
    ws = _get_sheet(wb)

    year = (audit_data or {}).get("year2023", {}) or {}

    if isinstance(year, dict) and "specifics" in year:
        year = dict(year)
        year.pop("specifics", None)

    _set_cell(ws, f"B{TITLE_ROWS['operational']}", "Activité opérationnelle")
    _set_cell(ws, f"B{TITLE_ROWS['buildings']}", "Bâtiments")
    _set_cell(ws, f"B{TITLE_ROWS['transport']}", "Transport")
    _set_cell(ws, f"B{TITLE_ROWS['utility']}", "Utilité")

    headers = year.get("utility_headers", {}) or {}
    _set_cell(ws, "G3", headers.get("util1_name", ""))
    _set_cell(ws, "G4", headers.get("util1_unit", ""))
    _set_cell(ws, "H3", headers.get("util2_name", ""))
    _set_cell(ws, "H4", headers.get("util2_unit", ""))

    col_map = {
        "name": "B", "electricity": "C", "gas": "D", "fuel": "E",
        "biogas": "F", "util1": "G", "util2": "H", "process": "I",
    }
    numeric_fields = {"electricity", "gas", "fuel", "biogas", "util1", "util2", "process"}

    for section_key in ["operational", "buildings", "transport", "utility"]:
        rows = year.get(section_key, []) or []
        start_row = SECTION_START_ROW[section_key]

        for r in range(MAX_ROWS_PER_SECTION):
            excel_row = start_row + r
            for field, col in col_map.items():
                _set_cell(ws, f"{col}{excel_row}", "" if field == "name" else None,
                          numeric=(field in numeric_fields))

        for idx, row in enumerate(rows[:MAX_ROWS_PER_SECTION]):
            excel_row = start_row + idx
            _set_cell(ws, f"B{excel_row}", row.get("name", ""))
            _set_cell(ws, f"C{excel_row}", row.get("electricity", None), numeric=True)
            _set_cell(ws, f"D{excel_row}", row.get("gas", None), numeric=True)
            _set_cell(ws, f"E{excel_row}", row.get("fuel", None), numeric=True)
            _set_cell(ws, f"F{excel_row}", row.get("biogas", None), numeric=True)
            _set_cell(ws, f"G{excel_row}", row.get("util1", None), numeric=True)
            _set_cell(ws, f"H{excel_row}", row.get("util2", None), numeric=True)
            _set_cell(ws, f"I{excel_row}", row.get("process", None), numeric=True)

    influence = year.get("influence_factors", []) or []
    for i in range(INFLUENCE_MAX_ROWS):
        excel_row = INFLUENCE_START_ROW + i
        _set_cell(ws, f"L{excel_row}", "")
        _set_cell(ws, f"M{excel_row}", None, numeric=True)
        _set_cell(ws, f"N{excel_row}", "")

    for i, row in enumerate(influence[:INFLUENCE_MAX_ROWS]):
        excel_row = INFLUENCE_START_ROW + i
        _set_cell(ws, f"L{excel_row}", row.get("description", ""))
        _set_cell(ws, f"M{excel_row}", row.get("value", None), numeric=True)
        _set_cell(ws, f"N{excel_row}", row.get("unit", ""))

    invoice = year.get("invoice_meter", {}) or {}
    _set_cell(ws, "C19", invoice.get("electricity", None), numeric=True)
    _set_cell(ws, "D19", invoice.get("gas", None), numeric=True)
    _set_cell(ws, "E19", invoice.get("fuel", None), numeric=True)
    _set_cell(ws, "F19", invoice.get("biogas", None), numeric=True)
    _set_cell(ws, "G19", invoice.get("util1", None), numeric=True)
    _set_cell(ws, "H19", invoice.get("util2", None), numeric=True)
    _set_cell(ws, "I19", invoice.get("process", None), numeric=True)

    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.save(excel_path)


def read_indices_from_excel(excel_path: Path) -> Dict[str, Any]:
    recalc_excel_in_place(excel_path)

    wb = load_workbook(excel_path, data_only=True, keep_links=False)
    ws = _get_sheet(wb)

    def clean(v: Any):
        if v is None:
            return None
        if isinstance(v, (int, float)):
            return v
        s = str(v).strip()
        if s.startswith("#"):  # Excel errors: #DIV/0!, #REF!, etc.
            return None
        s2 = s.replace(" ", "").replace(",", ".")
        try:
            return float(s2)
        except Exception:
            return s

    primary = {k: clean(ws[addr].value) for k, addr in INDICES_CELLS["primary"].items()}
    secondary = {k: clean(ws[addr].value) for k, addr in INDICES_CELLS["secondary"].items()}
    formulas_calculated = any(v is not None for v in list(primary.values()) + list(secondary.values()))
    return {
        "primary": primary,
        "secondary": secondary,
        "formulas_calculated": formulas_calculated,
    }


# ──────────────────────────────────────────────────────────────────────────────
# HELPERS AUDIT / ENERGY
# ──────────────────────────────────────────────────────────────────────────────

def _audit_to_data(audit: models.Audit) -> Dict[str, Any]:
    """Reconstruit le dict audit_data à partir de l'ORM Audit."""
    if audit is None:
        return {}
    energies = audit.energies or {}
    inf = audit.influence_factors or {}
    inv = audit.invoices or {}
    year_energy = energies.get("year2023", {})
    return {
        "year2023": {
            **year_energy,
            "influence_factors": inf.get("year2023", []),
            "invoice_meter": inv.get("year2023", {}),
        },
        "field_sources": audit.field_sources or {},
    }


def _upsert_audit(project_id: str, audit_data: Dict[str, Any], db: Session, field_sources: Dict[str, Any] = None) -> models.Audit:
    """Crée ou met à jour la ligne audits pour ce projet."""
    year2023 = (audit_data or {}).get("year2023", {}) or {}

    energies = {
        "year2023": {
            "utility_headers": year2023.get("utility_headers", {}),
            "operational":     year2023.get("operational", []),
            "buildings":       year2023.get("buildings", []),
            "transport":       year2023.get("transport", []),
            "utility":         year2023.get("utility", []),
        }
    }
    influence_factors = {"year2023": year2023.get("influence_factors", [])}
    invoices = {"year2023": year2023.get("invoice_meter", {})}

    audit = db.query(models.Audit).filter(models.Audit.project_id == project_id).first()
    if audit:
        audit.energies = energies
        audit.influence_factors = influence_factors
        audit.invoices = invoices
        flag_modified(audit, "energies")
        flag_modified(audit, "influence_factors")
        flag_modified(audit, "invoices")
        if field_sources:
            merged = {**(audit.field_sources or {}), **field_sources}
            audit.field_sources = merged
            flag_modified(audit, "field_sources")
    else:
        audit = models.Audit(
            id=str(uuid4()),
            project_id=project_id,
            energies=energies,
            influence_factors=influence_factors,
            invoices=invoices,
            field_sources=field_sources or {},
        )
        db.add(audit)
    db.commit()
    db.refresh(audit)
    return audit


def _reconstruct_energy(project_id: str, db: Session) -> Dict[str, Any]:
    """Reconstruit le dict energy_accounting à partir des lignes de la table."""
    records = db.query(models.EnergyRecord).filter(
        models.EnergyRecord.project_id == project_id
    ).all()
    years: Dict[str, Any] = {}
    for r in records:
        years[r.year] = {
            "year": r.year,
            "totals": {
                "electricity": r.electricity,
                "gas":         r.gas,
                "fuel":        r.fuel,
                "biogas":      r.biogas,
                "util1":       r.utility1,
                "util2":       r.utility2,
                "process":     r.process,
            },
            "details":      r.details or {},
            "notes":        r.notes or "",
            "field_sources": r.field_sources or {},
        }
    return {"years": years}


def _to_float(val) -> Optional[float]:
    if val is None or val == "":
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _upsert_energy_year(
    project_id: str, year_str: str, year_data: Dict[str, Any], db: Session
) -> None:
    """Crée ou met à jour une ligne energy_accounting pour une année."""
    totals = year_data.get("totals", {}) or {}
    details = year_data.get("details") or {}
    notes = year_data.get("notes", "") or ""

    incoming_fs = year_data.get("field_sources") or {}

    record = db.query(models.EnergyRecord).filter(
        models.EnergyRecord.project_id == project_id,
        models.EnergyRecord.year == year_str,
    ).first()

    if record:
        record.electricity = _to_float(totals.get("electricity"))
        record.gas         = _to_float(totals.get("gas"))
        record.fuel        = _to_float(totals.get("fuel"))
        record.biogas      = _to_float(totals.get("biogas"))
        record.utility1    = _to_float(totals.get("util1"))
        record.utility2    = _to_float(totals.get("util2"))
        record.process     = _to_float(totals.get("process"))
        record.notes       = notes
        record.details     = details
        flag_modified(record, "details")
        if incoming_fs:
            merged_fs = {**(record.field_sources or {}), **incoming_fs}
            record.field_sources = merged_fs
            flag_modified(record, "field_sources")
    else:
        record = models.EnergyRecord(
            id=str(uuid4()),
            project_id=project_id,
            year=year_str,
            electricity=_to_float(totals.get("electricity")),
            gas=_to_float(totals.get("gas")),
            fuel=_to_float(totals.get("fuel")),
            biogas=_to_float(totals.get("biogas")),
            utility1=_to_float(totals.get("util1")),
            utility2=_to_float(totals.get("util2")),
            process=_to_float(totals.get("process")),
            notes=notes,
            details=details,
            field_sources=incoming_fs or {},
        )
        db.add(record)


# ==============================
# ROUTES: PROJECTS CRUD
# ==============================
@app.get("/projects", response_model=List[schemas.Project])
def list_projects(db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    return db.query(models.Project).filter(models.Project.owner_id == current_user.id).all()


@app.post("/projects", response_model=schemas.Project)
def create_project(payload: schemas.ProjectCreate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    project_id = str(uuid4())
    excel_name = f"{project_id}.xlsx"
    excel_path = EXCEL_DIR / excel_name

    if not TEMPLATE_FILE.exists():
        raise HTTPException(status_code=500, detail="Excel template not found")

    copyfile(TEMPLATE_FILE, excel_path)

    project = models.Project(
        id=project_id,
        owner_id=current_user.id,
        created_at=datetime.now(timezone.utc).isoformat(),
        excel_file=excel_name,
        **payload.model_dump(),
    )
    db.add(project)
    db.commit()
    db.refresh(project)
    return project


@app.patch("/projects/{project_id}", response_model=schemas.Project)
def update_project(project_id: str, payload: schemas.ProjectUpdate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    project = db.query(models.Project).filter(models.Project.id == project_id, models.Project.owner_id == current_user.id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(project, field, value)

    db.commit()
    db.refresh(project)
    return project


@app.delete("/projects/{project_id}")
def delete_project(project_id: str, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    project = db.query(models.Project).filter(models.Project.id == project_id, models.Project.owner_id == current_user.id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    excel_path = EXCEL_DIR / project.excel_file
    if excel_path.exists():
        excel_path.unlink()

    db.delete(project)
    db.commit()
    return {"status": "deleted", "id": project_id}


# ==============================
# ROUTES: AUDIT + EXCEL
# ==============================
@app.get("/projects/{project_id}/audit")
def get_project_audit(project_id: str, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    project = db.query(models.Project).filter(models.Project.id == project_id, models.Project.owner_id == current_user.id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    audit = db.query(models.Audit).filter(models.Audit.project_id == project_id).first()
    return _audit_to_data(audit)


@app.get("/projects/{project_id}/audit/energie-chauffage")
def get_audit_energie_chauffage(
    project_id: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """Retourne la consommation totale de chauffage (kWh/an) issue de l'audit :
    somme des colonnes electricity + gas + fuel sur toutes les sections du template."""
    project = db.query(models.Project).filter(
        models.Project.id == project_id,
        models.Project.owner_id == current_user.id,
    ).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    audit = db.query(models.Audit).filter(models.Audit.project_id == project_id).first()
    if not audit or not audit.energies:
        return {"consommation_kwh_an": None}

    year = (audit.energies or {}).get("year2023", {}) or {}
    total = 0.0
    has_any = False
    for section_key in ["operational", "buildings", "transport", "utility"]:
        for row in (year.get(section_key, []) or []):
            for field in ("electricity", "gas", "fuel"):
                v = row.get(field)
                if v is not None:
                    try:
                        total += float(v)
                        has_any = True
                    except (TypeError, ValueError):
                        pass

    return {"consommation_kwh_an": round(total, 2) if has_any else None}


@app.patch("/projects/{project_id}/audit")
def update_project_audit(project_id: str, payload: schemas.AuditUpdate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    project = db.query(models.Project).filter(models.Project.id == project_id, models.Project.owner_id == current_user.id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    audit = _upsert_audit(project_id, payload.audit_data, db, field_sources=payload.field_sources)
    audit_data = _audit_to_data(audit)

    write_audit_to_excel(project, audit_data)

    return {"status": "ok", "audit_data": audit_data}


@app.get("/projects/{project_id}/excel")
def download_project_excel(project_id: str, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    project = db.query(models.Project).filter(models.Project.id == project_id, models.Project.owner_id == current_user.id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    _ensure_excel(project, db)
    excel_path = EXCEL_DIR / project.excel_file

    return FileResponse(
        path=str(excel_path),
        filename=f"{_safe_filename(project.project_name)}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ==============================
# ROUTES: DOCUMENT EXTRACTION
# ==============================
@app.post("/projects/{project_id}/extract-document")
async def extract_document(
    project_id: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    project = db.query(models.Project).filter(
        models.Project.id == project_id,
        models.Project.owner_id == current_user.id,
    ).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    content_type = file.content_type or ""
    if content_type not in ("application/pdf", "image/jpeg", "image/png"):
        raise HTTPException(
            status_code=400,
            detail="Type de fichier non supporté. Utilisez PDF, JPEG ou PNG.",
        )

    file_bytes = await file.read()
    b64_data = base64.standard_b64encode(file_bytes).decode("utf-8")

    if content_type == "application/pdf":
        content_block = {
            "type": "document",
            "source": {"type": "base64", "media_type": "application/pdf", "data": b64_data},
        }
    else:
        content_block = {
            "type": "image",
            "source": {"type": "base64", "media_type": content_type, "data": b64_data},
        }

    prompt = (
        "Tu es un assistant spécialisé en audit énergétique belge. "
        "Analyse ce document et extrait les informations suivantes :\n"
        "- type d'énergie (electricite/gaz/fuel/biogas)\n"
        "- consommation (nombre)\n"
        "- unite (kWh/litres/m3)\n"
        "- annee (nombre)\n"
        "- cout_total (nombre en euros)\n"
        "Retourne UNIQUEMENT un JSON valide sans markdown."
    )

    try:
        client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))
        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=1024,
            messages=[
                {
                    "role": "user",
                    "content": [
                        content_block,
                        {"type": "text", "text": prompt},
                    ],
                }
            ],
        )
        raw = message.content[0].text.strip()
        extracted = json.loads(raw)
    except json.JSONDecodeError:
        raise HTTPException(status_code=422, detail="Claude n'a pas retourné un JSON valide.")
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    return extracted


# ==============================
# ROUTES: INDICES
# ==============================
@app.get("/projects/{project_id}/indices")
def get_indices(project_id: str, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    project = db.query(models.Project).filter(models.Project.id == project_id, models.Project.owner_id == current_user.id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    _ensure_excel(project, db)
    excel_path = EXCEL_DIR / project.excel_file

    return read_indices_from_excel(excel_path)


# ==============================
# ROUTES: ENERGY ACCOUNTING
# ==============================
@app.get("/projects/{project_id}/energy-accounting")
def get_energy_accounting(project_id: str, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    project = db.query(models.Project).filter(models.Project.id == project_id, models.Project.owner_id == current_user.id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    return _reconstruct_energy(project_id, db)


@app.patch("/projects/{project_id}/energy-accounting")
def update_energy_accounting(project_id: str, payload: schemas.EnergyAccountingUpdate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    project = db.query(models.Project).filter(models.Project.id == project_id, models.Project.owner_id == current_user.id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    years = (payload.energy_accounting or {}).get("years", {}) or {}
    for year_str, year_data in years.items():
        _upsert_energy_year(project_id, year_str, year_data, db)

    db.commit()
    return {"status": "ok", "energy_accounting": _reconstruct_energy(project_id, db)}


@app.post("/projects/{project_id}/energy-accounting/import-from-audit")
def import_energy_from_audit(project_id: str, payload: schemas.EnergyYearImportRequest, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    project = db.query(models.Project).filter(models.Project.id == project_id, models.Project.owner_id == current_user.id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    audit = db.query(models.Audit).filter(models.Audit.project_id == project_id).first()
    audit_data = _audit_to_data(audit)
    year2023_data = (audit_data or {}).get("year2023", {}) or {}
    year = str(payload.year)

    def sum_field(rows, field):
        total = 0.0
        for r in rows or []:
            v = r.get(field, None)
            if v is None or str(v).strip() == "":
                continue
            try:
                total += float(str(v).replace(" ", "").replace(",", "."))
            except Exception:
                pass
        return total

    sections = ["operational", "buildings", "transport", "utility"]
    all_rows = []
    for sk in sections:
        all_rows += (year2023_data.get(sk, []) or [])

    year_data = {
        "year": year,
        "totals": {
            "electricity": sum_field(all_rows, "electricity"),
            "gas":         sum_field(all_rows, "gas"),
            "fuel":        sum_field(all_rows, "fuel"),
            "biogas":      sum_field(all_rows, "biogas"),
            "util1":       sum_field(all_rows, "util1"),
            "util2":       sum_field(all_rows, "util2"),
            "process":     sum_field(all_rows, "process"),
        },
        "details": {sk: year2023_data.get(sk, []) for sk in sections},
        "notes": "",
    }

    _upsert_energy_year(project_id, year, year_data, db)
    db.commit()
    return {"status": "ok", "energy_accounting": _reconstruct_energy(project_id, db)}


# ==============================
# ROUTES: REPORT
# ==============================
@app.get("/projects/{project_id}/report")
def get_project_report(project_id: str, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    project = db.query(models.Project).filter(models.Project.id == project_id, models.Project.owner_id == current_user.id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    report = db.query(models.Report).filter(models.Report.project_id == project_id).first()
    if not report:
        return {}

    return {
        "audit_type":       report.audit_type or "",
        "audit_theme":      report.theme or "",
        "provider_company": report.provider_name or "",
        "auditor_name":     report.auditor_name or "",
        "amureba_skills":   report.competences or "",
        "field_sources":    report.field_sources or {},
    }


@app.patch("/projects/{project_id}/report")
def update_project_report(project_id: str, payload: schemas.ReportUpdate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    project = db.query(models.Project).filter(models.Project.id == project_id, models.Project.owner_id == current_user.id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    data = payload.report_data or {}
    report = db.query(models.Report).filter(models.Report.project_id == project_id).first()

    if report:
        report.audit_type    = data.get("audit_type", report.audit_type)
        report.theme         = data.get("audit_theme", report.theme)
        report.provider_name = data.get("provider_company", report.provider_name)
        report.auditor_name  = data.get("auditor_name", report.auditor_name)
        report.competences   = data.get("amureba_skills", report.competences)
    else:
        report = models.Report(
            id=str(uuid4()),
            project_id=project_id,
            audit_type    = data.get("audit_type"),
            theme         = data.get("audit_theme"),
            provider_name = data.get("provider_company"),
            auditor_name  = data.get("auditor_name"),
            competences   = data.get("amureba_skills"),
        )
        db.add(report)

    if payload.field_sources:
        merged_fs = {**(report.field_sources or {}), **payload.field_sources}
        report.field_sources = merged_fs

    db.commit()
    db.refresh(report)

    report_data = {
        "audit_type":       report.audit_type or "",
        "audit_theme":      report.theme or "",
        "provider_company": report.provider_name or "",
        "auditor_name":     report.auditor_name or "",
        "amureba_skills":   report.competences or "",
        "field_sources":    report.field_sources or {},
    }
    return {"status": "ok", "report_data": report_data}


@app.get("/projects/{project_id}/report/docx")
def download_project_report_docx(project_id: str, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    project = db.query(models.Project).filter(models.Project.id == project_id, models.Project.owner_id == current_user.id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    if not REPORT_TEMPLATE_FILE.exists():
        raise HTTPException(status_code=500, detail="Report template not found")

    report = db.query(models.Report).filter(models.Report.project_id == project_id).first()
    audit_type = (report.audit_type if report else None) or project.audit_type or ""
    audit_type = audit_type.strip()

    context = {
        "audit_type":        audit_type,
        "audit_theme":       (report.theme          if report else "") or "",
        "provider_company":  (report.provider_name  if report else "") or "",
        "auditor_name":      (report.auditor_name   if report else "") or "",
        "amureba_skills":    (report.competences    if report else "") or "",
        "audit_global_box":  "☑" if audit_type.lower() == "audit global"  else "☐",
        "audit_partiel_box": "☑" if audit_type.lower() == "audit partiel" else "☐",
    }

    out_path = REPORT_DIR / f"{project.id}_report.docx"
    doc = DocxTemplate(str(REPORT_TEMPLATE_FILE))
    doc.render(context)
    doc.save(str(out_path))

    return FileResponse(
        path=str(out_path),
        filename=f"{_safe_filename(project.project_name)}_rapport.docx",
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    )


# ==============================
# ROUTES: EVENTS (Agenda)
# ==============================
@app.get("/events", response_model=List[schemas.EventSchema])
def list_events(project_id: Optional[str] = None, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    q = db.query(models.Event)
    if project_id:
        q = q.filter(models.Event.project_id == project_id)
    return q.order_by(models.Event.start).all()


@app.post("/events", response_model=schemas.EventSchema)
def create_event(payload: schemas.EventCreate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    event = models.Event(id=str(uuid4()), **payload.model_dump())
    db.add(event)
    db.commit()
    db.refresh(event)
    return event


@app.patch("/events/{event_id}", response_model=schemas.EventSchema)
def update_event(event_id: str, payload: schemas.EventCreate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    event = db.query(models.Event).filter(models.Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(event, field, value)

    db.commit()
    db.refresh(event)
    return event


@app.delete("/events/{event_id}")
def delete_event(event_id: str, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    event = db.query(models.Event).filter(models.Event.id == event_id).first()
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    db.delete(event)
    db.commit()
    return {"status": "deleted", "id": event_id}


# ==============================
# ROUTES: CLIENT REQUESTS
# ==============================
@app.get("/client-requests", response_model=List[schemas.ClientRequestSchema])
def list_client_requests(project_id: Optional[str] = None, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    q = db.query(models.ClientRequest)
    if project_id:
        q = q.filter(models.ClientRequest.project_id == project_id)
    return q.all()


@app.post("/client-requests", response_model=schemas.ClientRequestSchema)
def create_client_request(payload: schemas.ClientRequestCreate, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    cr = models.ClientRequest(id=str(uuid4()), **payload.model_dump())
    db.add(cr)
    db.commit()
    db.refresh(cr)
    return cr


@app.patch("/client-requests/{request_id}", response_model=schemas.ClientRequestSchema)
def update_client_request(request_id: str, payload: schemas.ClientRequestPatch, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    cr = db.query(models.ClientRequest).filter(models.ClientRequest.id == request_id).first()
    if not cr:
        raise HTTPException(status_code=404, detail="Client request not found")

    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(cr, field, value)
        if field in ("documents", "received_files"):
            flag_modified(cr, field)

    db.commit()
    db.refresh(cr)
    return cr


@app.delete("/client-requests/{request_id}")
def delete_client_request(request_id: str, db: Session = Depends(get_db), current_user: models.User = Depends(get_current_user)):
    cr = db.query(models.ClientRequest).filter(models.ClientRequest.id == request_id).first()
    if not cr:
        raise HTTPException(status_code=404, detail="Client request not found")

    db.delete(cr)
    db.commit()
    return {"status": "deleted", "id": request_id}


# ==============================
# ROUTES: PROJECT DOCUMENTS
# ==============================

_ANALYZE_PROMPT = (
    "Tu es un assistant spécialisé en audit énergétique belge. "
    "Analyse ce document et extrait TOUTES les informations pertinentes. "
    "Retourne UNIQUEMENT un JSON valide sans markdown avec cette structure :\n"
    '{"type_document":"facture_electricite|facture_gaz|facture_fuel|releve|contrat|autre",'
    '"energie":"electricite|gaz|fuel|biogas|null",'
    '"annee":nombre_ou_null,'
    '"consommation":nombre_ou_null,'
    '"unite":"kWh|m3|litres|null",'
    '"cout_total":nombre_ou_null,'
    '"periode_debut":"YYYY-MM ou null",'
    '"periode_fin":"YYYY-MM ou null",'
    '"fournisseur":"string ou null",'
    '"adresse_site":"string ou null",'
    '"nom_client":"string ou null",'
    '"type_batiment":"string ou null",'
    '"surface":nombre_ou_null,'
    '"auditeur":"string ou null",'
    '"notes":"string ou null"}'
)


def _doc_to_dict(doc: models.ProjectDocument) -> Dict[str, Any]:
    return {
        "id": doc.id,
        "project_id": doc.project_id,
        "owner_id": doc.owner_id,
        "filename": doc.filename,
        "original_name": doc.original_name,
        "file_type": doc.file_type,
        "doc_type": doc.doc_type,
        "status": doc.status,
        "extracted_data": doc.extracted_data,
        "created_at": doc.created_at,
    }


def _analyze_one(doc: models.ProjectDocument) -> None:
    """Analyse un document avec Claude et met à jour doc.status / doc.extracted_data (sans commit)."""
    b64 = base64.standard_b64encode(doc.file_data).decode("utf-8")
    if doc.file_type == "application/pdf":
        content_block = {
            "type": "document",
            "source": {"type": "base64", "media_type": "application/pdf", "data": b64},
        }
    else:
        content_block = {
            "type": "image",
            "source": {"type": "base64", "media_type": doc.file_type, "data": b64},
        }
    print(f"[analyze] envoi à Claude, media_type: {doc.file_type}, doc_id: {doc.id}", flush=True)
    try:
        client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))
        message = client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=1024,
            messages=[{
                "role": "user",
                "content": [content_block, {"type": "text", "text": _ANALYZE_PROMPT}],
            }],
        )
        print(f"[analyze] réponse brute: {message.content}", flush=True)
        raw = message.content[0].text.strip()
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        raw = raw.strip()
        doc.extracted_data = json.loads(raw)
        doc.status = "analyzed"
        flag_modified(doc, "extracted_data")
    except json.JSONDecodeError as e:
        print(f"[analyze] ERREUR JSONDecodeError: {str(e)}", flush=True)
        traceback.print_exc()
        doc.status = "error"
    except Exception as e:
        print(f"[analyze] ERREUR: {str(e)}", flush=True)
        traceback.print_exc()
        doc.status = "error"


@app.post("/projects/{project_id}/documents")
async def upload_document(
    project_id: str,
    file: UploadFile = File(...),
    doc_type: str = Form("autre"),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    project = db.query(models.Project).filter(
        models.Project.id == project_id,
        models.Project.owner_id == current_user.id,
    ).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    content_type = file.content_type or ""
    if content_type not in ("application/pdf", "image/jpeg", "image/png"):
        raise HTTPException(status_code=400, detail="Type non supporté. Utilisez PDF, JPEG ou PNG.")

    file_bytes = await file.read()
    doc = models.ProjectDocument(
        id=str(uuid4()),
        project_id=project_id,
        owner_id=current_user.id,
        filename=file.filename or "document",
        original_name=file.filename or "document",
        file_type=content_type,
        doc_type=doc_type,
        file_data=file_bytes,
        status="pending",
        created_at=datetime.now(timezone.utc).isoformat(),
    )
    db.add(doc)
    db.commit()
    db.refresh(doc)
    return _doc_to_dict(doc)


@app.get("/projects/{project_id}/documents")
def list_documents(
    project_id: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    project = db.query(models.Project).filter(
        models.Project.id == project_id,
        models.Project.owner_id == current_user.id,
    ).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    docs = db.query(models.ProjectDocument).filter(
        models.ProjectDocument.project_id == project_id
    ).order_by(models.ProjectDocument.created_at.desc()).all()
    return [_doc_to_dict(d) for d in docs]


@app.delete("/projects/{project_id}/documents/{doc_id}")
def delete_document(
    project_id: str,
    doc_id: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    project = db.query(models.Project).filter(
        models.Project.id == project_id,
        models.Project.owner_id == current_user.id,
    ).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    doc = db.query(models.ProjectDocument).filter(
        models.ProjectDocument.id == doc_id,
        models.ProjectDocument.project_id == project_id,
    ).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    db.delete(doc)
    db.commit()
    return {"status": "deleted", "id": doc_id}


@app.post("/projects/{project_id}/documents/analyze-all")
def analyze_all_documents(
    project_id: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    project = db.query(models.Project).filter(
        models.Project.id == project_id,
        models.Project.owner_id == current_user.id,
    ).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    docs = db.query(models.ProjectDocument).filter(
        models.ProjectDocument.project_id == project_id,
        models.ProjectDocument.status.in_(["pending", "error"]),
    ).all()

    for doc in docs:
        _analyze_one(doc)

    db.commit()
    return {"analyzed": len(docs), "documents": [_doc_to_dict(d) for d in docs]}


@app.get("/projects/{project_id}/documents/{doc_id}/file")
def get_document_file(
    project_id: str,
    doc_id: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    project = db.query(models.Project).filter(
        models.Project.id == project_id,
        models.Project.owner_id == current_user.id,
    ).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    doc = db.query(models.ProjectDocument).filter(
        models.ProjectDocument.id == doc_id,
        models.ProjectDocument.project_id == project_id,
    ).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    return Response(content=bytes(doc.file_data), media_type=doc.file_type)


@app.post("/projects/{project_id}/documents/{doc_id}/analyze")
def analyze_document(
    project_id: str,
    doc_id: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    project = db.query(models.Project).filter(
        models.Project.id == project_id,
        models.Project.owner_id == current_user.id,
    ).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    doc = db.query(models.ProjectDocument).filter(
        models.ProjectDocument.id == doc_id,
        models.ProjectDocument.project_id == project_id,
    ).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Document not found")

    _analyze_one(doc)
    db.commit()
    db.refresh(doc)
    return _doc_to_dict(doc)


# ==============================
# ROUTES: PLAN D'AMÉLIORATION
# ==============================

@app.get("/projects/{project_id}/improvement-actions", response_model=List[schemas.ImprovementActionSchema])
def list_improvement_actions(
    project_id: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    project = db.query(models.Project).filter(
        models.Project.id == project_id,
        models.Project.owner_id == current_user.id,
    ).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    return (
        db.query(models.ImprovementAction)
        .filter(models.ImprovementAction.project_id == project_id)
        .order_by(models.ImprovementAction.created_at)
        .all()
    )


@app.post("/projects/{project_id}/improvement-actions", response_model=schemas.ImprovementActionSchema, status_code=201)
def create_improvement_action(
    project_id: str,
    payload: schemas.ImprovementActionCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    project = db.query(models.Project).filter(
        models.Project.id == project_id,
        models.Project.owner_id == current_user.id,
    ).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    action = models.ImprovementAction(
        id=str(uuid4()),
        project_id=project_id,
        owner_id=current_user.id,
        created_at=datetime.now(timezone.utc).isoformat(),
        **payload.model_dump(),
    )
    db.add(action)
    db.commit()
    db.refresh(action)
    return action


@app.put("/projects/{project_id}/improvement-actions/{action_id}", response_model=schemas.ImprovementActionSchema)
def update_improvement_action(
    project_id: str,
    action_id: str,
    payload: schemas.ImprovementActionUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    project = db.query(models.Project).filter(
        models.Project.id == project_id,
        models.Project.owner_id == current_user.id,
    ).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    action = db.query(models.ImprovementAction).filter(
        models.ImprovementAction.id == action_id,
        models.ImprovementAction.project_id == project_id,
    ).first()
    if not action:
        raise HTTPException(status_code=404, detail="Action not found")
    for field, value in payload.model_dump(exclude_unset=True).items():
        setattr(action, field, value)
    db.commit()
    db.refresh(action)
    return action


@app.delete("/projects/{project_id}/improvement-actions/{action_id}", status_code=204)
def delete_improvement_action(
    project_id: str,
    action_id: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    project = db.query(models.Project).filter(
        models.Project.id == project_id,
        models.Project.owner_id == current_user.id,
    ).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    action = db.query(models.ImprovementAction).filter(
        models.ImprovementAction.id == action_id,
        models.ImprovementAction.project_id == project_id,
    ).first()
    if not action:
        raise HTTPException(status_code=404, detail="Action not found")
    db.delete(action)
    db.commit()


# ──────────────────────────────────────────────────────────────────────────────
# HELPERS — Plan d'amélioration Excel
# ──────────────────────────────────────────────────────────────────────────────

_IMPROVEMENT_TYPE_MAP = [
    (["ser ", "pv", "éolien", "eolien", "géothermie", "geothermie", "solaire"], "SER_PV"),
    (["electrif", "électrif"], "ELECTRIFICATION"),
    (["efficac"], "EFFICACITE_ENERGETIQUE"),
    (["ccu"], "CCU"),
    (["ppa"], "PPA"),
    (["fluide", "frigorif"], "FLUIDE_FRIGORIGENE"),
]


def _normalize_improvement_type(raw: str) -> Optional[str]:
    if not raw:
        return None
    lower = raw.lower()
    for keywords, enum_val in _IMPROVEMENT_TYPE_MAP:
        if any(k in lower for k in keywords):
            return enum_val
    return raw  # keep raw string if nothing matched


def _parse_aa_sheet(ws) -> Optional[Dict[str, Any]]:
    """Extrait les données d'une feuille AAx. Retourne None si la feuille est vide."""
    intitule = ws["B9"].value
    if not intitule or not str(intitule).strip():
        return None

    def _oui_non(v) -> Optional[bool]:
        if v is None:
            return None
        return str(v).strip().upper() == "OUI"

    # Conditions préalables (C31, C33, C35)
    conds = []
    for r in [31, 33, 35]:
        v = ws[f"C{r}"].value
        if v and str(v).strip() and str(v).strip().upper() not in ("NA", "N/A", ""):
            conds.append(str(v).strip())

    ref = ws["F5"].value
    type_raw = ws["B13"].value
    classif = ws["F27"].value
    duree_raw = _to_number(ws["K18"].value)

    return {
        "reference": str(ref).strip() if ref else None,
        "intitule": str(intitule).strip(),
        "type_amelioration": _normalize_improvement_type(str(type_raw) if type_raw else ""),
        "classification": str(classif).strip() if classif else None,
        "conditions_prealables": "\n".join(conds) if conds else None,
        "investissement": _to_number(ws["G61"].value),
        "economie_energie": _to_number(ws["G77"].value),
        "economie_co2": _to_number(ws["G87"].value),
        "duree_amortissement": int(duree_raw) if duree_raw is not None else None,
        "irr_avant_impot": _to_number(ws["N10"].value),
        "pbt_avant_impot": _to_number(ws["N15"].value),
        "irr_apres_impot": _to_number(ws["N17"].value),
        "pbt_apres_impot": _to_number(ws["N18"].value),
        "entreprise_ets": _oui_non(ws["K22"].value),
        "deduction_fiscale": _oui_non(ws["K23"].value),
        "description": None,
        "situation_existante": None,
    }


# ──────────────────────────────────────────────────────────────────────────────
# ROUTES — Export / Import / Prefill Excel AMUREBA
# ──────────────────────────────────────────────────────────────────────────────

_PREFILL_SYSTEM = """\
Tu es un expert en audit énergétique industriel (méthode AMUREBA belge).
Tu reçois des données extraites de factures énergétiques d'un bâtiment/entreprise.
Tu dois proposer des actions d'amélioration énergétique concrètes et chiffrées,
adaptées aux consommations détectées, en JSON uniquement — aucun texte autour.
"""

_PREFILL_USER_TPL = """\
Données extraites des documents du projet "{project_name}" :

{extracted_json}

Génère entre 1 et 5 actions d'amélioration énergétique réalistes.
Réponds UNIQUEMENT avec un tableau JSON de la forme :
[
  {{
    "intitule": "...",
    "type_amelioration": "...",   // ex: Efficacité Energétique, SER PV, Electrification…
    "classification": "A",        // A ou B
    "investissement_k_eur": 0,    // en k€ (nombre)
    "economie_energie_mwh_an": 0, // MWh/an
    "economie_co2_kg_an": 0,      // kg CO2/an
    "duree_amortissement": 8,     // années
    "ets": "NON",                 // OUI ou NON
    "deduction_fiscale": "OUI",   // OUI ou NON
    "situation_existante": "...",
    "description": "...",
    "sources": {{
      "investissement_k_eur":    {{"document": "<nom_fichier ou null>", "field": "<champ_source ou null>", "estimated": true}},
      "economie_energie_mwh_an": {{"document": "<nom_fichier ou null>", "field": "<champ_source ou null>", "estimated": false}},
      "economie_co2_kg_an":      {{"document": "<nom_fichier ou null>", "field": "<champ_source ou null>", "estimated": true}},
      "classification":          {{"document": "<nom_fichier ou null>", "field": "<champ_source ou null>", "estimated": true}},
      "type_amelioration":       {{"document": "<nom_fichier ou null>", "field": "<champ_source ou null>", "estimated": true}}
    }}
  }}
]
Règles pour "sources" :
- Si la valeur est directement issue d'un document fourni : "document" = nom exact du fichier,
  "field" = nom du champ extrait utilisé (ex: "consommation_electricite"), "estimated": false.
- Si la valeur est estimée / calculée par l'IA sans source directe : "document": null, "field": null, "estimated": true.
Si les données sont insuffisantes pour chiffrer, utilise 0 pour les valeurs numériques
et indique des ordres de grandeur raisonnables basés sur le secteur.
"""


async def _get_prefill_actions(
    project_id: str,
    db: Session,
    current_user: models.User,
) -> tuple:
    """
    Shared helper: fetch extracted_data from project documents, call Claude,
    return (project, entity_name, actions_list).
    Raises HTTPException on error.
    """
    project = db.query(models.Project).filter(
        models.Project.id == project_id,
        models.Project.owner_id == current_user.id,
    ).first()
    if not project:
        raise HTTPException(status_code=404, detail="Projet introuvable")

    docs = db.query(models.ProjectDocument).filter(
        models.ProjectDocument.project_id == project_id,
        models.ProjectDocument.extracted_data.isnot(None),
    ).all()

    extracted_parts = [
        {"document": d.original_name, "type": d.doc_type, "data": d.extracted_data}
        for d in docs if d.extracted_data
    ]

    if not extracted_parts:
        raise HTTPException(
            status_code=422,
            detail=(
                "Aucun document analysé trouvé pour ce projet. "
                "Uploadez et analysez des factures dans le module Documents d'abord."
            ),
        )

    entity_name = project.client_name or project.project_name or ""
    extracted_json = json.dumps(extracted_parts, ensure_ascii=False, indent=2)
    user_msg = _PREFILL_USER_TPL.format(
        project_name=entity_name,
        extracted_json=extracted_json,
    )

    try:
        claude_client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))
        msg = claude_client.messages.create(
            model="claude-sonnet-4-20250514",
            max_tokens=4096,
            system=_PREFILL_SYSTEM,
            messages=[{"role": "user", "content": user_msg}],
        )
        raw_text = msg.content[0].text.strip()
        json_match = re.search(r"\[.*\]", raw_text, re.DOTALL)
        if not json_match:
            raise ValueError("Aucun tableau JSON trouvé dans la réponse Claude")
        actions: List[Dict[str, Any]] = json.loads(json_match.group(0))
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Erreur lors de l'appel à Claude : {e}")

    return project, entity_name, actions



@app.post("/projects/{project_id}/improvement-actions/prefill-preview")
async def prefill_preview(
    project_id: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """
    Appelle Claude et retourne le JSON des actions proposées sans générer de fichier.
    Utilisé pour afficher un aperçu avant téléchargement.
    """
    _, entity_name, actions = await _get_prefill_actions(project_id, db, current_user)
    return {
        "entity_name": entity_name,
        "nb_actions": len(actions),
        "actions": [
            {
                "sheet": f"AA{i + 1}",
                "intitule": a.get("intitule"),
                "type_amelioration": a.get("type_amelioration"),
                "classification": a.get("classification"),
                "investissement_k_eur": a.get("investissement_k_eur"),
                "economie_energie_mwh_an": a.get("economie_energie_mwh_an"),
                "economie_co2_kg_an": a.get("economie_co2_kg_an"),
                "duree_amortissement": a.get("duree_amortissement"),
                # sources: dict {field: {document, field, estimated}} — peut être absent si Claude ne le fournit pas
                "sources": a.get("sources") or {},
            }
            for i, a in enumerate(actions[:9])
        ],
    }


@app.post("/projects/{project_id}/improvement-actions/prefill-excel")
async def prefill_improvement_actions_excel(
    project_id: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """
    Appelle Claude, pré-remplit le template AMUREBA, persiste en base et retourne le .xlsx.
    """
    project, entity_name, actions = await _get_prefill_actions(project_id, db, current_user)

    if not TEMPLATE_FILE.exists():
        raise HTTPException(status_code=500, detail="Template AMUREBA introuvable")
    try:
        file_bytes = _write_template_zipfile(TEMPLATE_FILE, entity_name, actions)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur lors de la génération du fichier: {e}")

    # ── Persistance ───────────────────────────────────────────────────────────
    now = datetime.now(timezone.utc).isoformat()
    summary = {
        "entity_name": entity_name,
        "nb_actions": len(actions),
        "actions": [
            {
                "sheet": f"AA{i + 1}",
                "intitule": a.get("intitule"),
                "type_amelioration": a.get("type_amelioration"),
                "classification": a.get("classification"),
                "investissement_k_eur": a.get("investissement_k_eur"),
                "economie_energie_mwh_an": a.get("economie_energie_mwh_an"),
                "economie_co2_kg_an": a.get("economie_co2_kg_an"),
                "duree_amortissement": a.get("duree_amortissement"),
                "sources": a.get("sources") or {},
            }
            for i, a in enumerate(actions[:9])
        ],
    }
    project.prefill_summary = summary
    project.prefilled_excel = file_bytes
    project.prefilled_at = now
    flag_modified(project, "prefill_summary")
    db.commit()

    safe_name = _safe_filename(entity_name or project.project_name)
    return StreamingResponse(
        io.BytesIO(file_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="AMUREBA_prefill_{safe_name}.xlsx"'},
    )


@app.get("/projects/{project_id}/improvement-actions/prefill-status")
def get_prefill_status(
    project_id: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """Retourne le statut du pré-remplissage IA pour ce projet."""
    project = db.query(models.Project).filter(
        models.Project.id == project_id,
        models.Project.owner_id == current_user.id,
    ).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    return {
        "has_prefilled_excel": project.prefilled_excel is not None,
        "prefilled_at": project.prefilled_at,
        "prefill_summary": project.prefill_summary,
    }


@app.get("/projects/{project_id}/improvement-actions/export-excel")
def export_improvement_actions_excel(
    project_id: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """
    Retourne le fichier xlsx pré-rempli par l'IA s'il existe,
    sinon le template vierge avec le nom de l'entité.
    """
    project = db.query(models.Project).filter(
        models.Project.id == project_id,
        models.Project.owner_id == current_user.id,
    ).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    safe_name = _safe_filename(project.project_name)

    # Retourner le fichier pré-rempli s'il est sauvegardé
    if project.prefilled_excel:
        return StreamingResponse(
            io.BytesIO(project.prefilled_excel),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="AMUREBA_prefill_{safe_name}.xlsx"'},
        )

    # Sinon : template vierge avec le nom de l'entité
    if not TEMPLATE_FILE.exists():
        raise HTTPException(status_code=500, detail="Template AMUREBA introuvable côté serveur")

    entity_name = project.client_name or project.project_name or ""
    try:
        file_bytes = _write_template_zipfile(TEMPLATE_FILE, entity_name, [])
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur lors de la génération du fichier: {e}")

    return StreamingResponse(
        io.BytesIO(file_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="AMUREBA_{safe_name}.xlsx"'},
    )


@app.post("/projects/{project_id}/improvement-actions/apply-prefill")
async def apply_prefill(
    project_id: str,
    payload: schemas.ApplyPrefillRequest,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """
    Reçoit la liste des changements sélectionnés/rejetés par l'auditeur.
    - Injecte uniquement les items sélectionnés dans le xlsx.
    - Sauvegarde le fichier + tous les items (avec statut) en base + historique.
    """
    project = db.query(models.Project).filter(
        models.Project.id == project_id,
        models.Project.owner_id == current_user.id,
    ).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    if not TEMPLATE_FILE.exists():
        raise HTTPException(status_code=500, detail="Template AMUREBA introuvable")

    entity_name = project.client_name or project.project_name or ""
    actions_by_sheet: Dict[str, Dict[str, Any]] = {}
    for item in payload.changes:
        if not item.selected:
            continue

        action = actions_by_sheet.setdefault(item.sheet, {"sheet": item.sheet})
        action[item.field] = item.value

    selected_actions = [
        actions_by_sheet[sheet]
        for sheet in sorted(actions_by_sheet.keys(), key=lambda s: int(re.sub(r"[^0-9]", "", s) or 0))
    ]

    print("[DEBUG] apply_prefill selected_actions:", flush=True)
    for action in selected_actions:
        print(f"  - {action}", flush=True)

    debug_sheet_changes = _build_prefill_sheet_changes(entity_name, selected_actions)
    print("[DEBUG] apply_prefill sheet_changes:", flush=True)
    for sheet_name, changes in debug_sheet_changes.items():
        print(f"  - {sheet_name}: {changes}", flush=True)

    try:
        file_bytes = _write_template_zipfile(TEMPLATE_FILE, entity_name, selected_actions)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur génération xlsx : {e}")

    now = datetime.now(timezone.utc).isoformat()
    all_items = [item.model_dump() for item in payload.changes]
    summary = {"items": all_items}

    project.prefilled_excel = file_bytes
    project.prefill_summary = summary
    project.prefilled_at = now
    flag_modified(project, "prefill_summary")

    db.add(models.PlanAmeliorationHistory(
        id=str(uuid4()),
        project_id=project_id,
        owner_id=current_user.id,
        action_type="AI_PREFILL",
        changes=summary,
        created_at=now,
    ))
    db.commit()

    safe_name = _safe_filename(project.project_name)
    return StreamingResponse(
        io.BytesIO(file_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="AMUREBA_prefill_{safe_name}.xlsx"'},
    )


@app.get("/projects/{project_id}/improvement-actions/history")
def get_prefill_history(
    project_id: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """Retourne l'historique complet des modifications du plan d'amélioration."""
    project = db.query(models.Project).filter(
        models.Project.id == project_id,
        models.Project.owner_id == current_user.id,
    ).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    entries = (
        db.query(models.PlanAmeliorationHistory)
        .filter(models.PlanAmeliorationHistory.project_id == project_id)
        .order_by(models.PlanAmeliorationHistory.created_at.desc())
        .all()
    )
    return [
        {
            "id": e.id,
            "action_type": e.action_type,
            "changes": e.changes,
            "created_at": e.created_at,
        }
        for e in entries
    ]


@app.post("/projects/{project_id}/improvement-actions/import-excel")
async def import_improvement_actions_excel(
    project_id: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """Lit un Excel AMUREBA complété et upsert les actions dans la base."""
    project = db.query(models.Project).filter(
        models.Project.id == project_id,
        models.Project.owner_id == current_user.id,
    ).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True, keep_links=False)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Fichier Excel invalide: {e}")

    parsed: List[Dict[str, Any]] = []
    for i in range(1, 10):
        sheet_name = f"AA{i}"
        if sheet_name not in wb.sheetnames:
            continue
        try:
            data = _parse_aa_sheet(wb[sheet_name])
        except Exception:
            continue
        if data:
            parsed.append(data)

    if not parsed:
        raise HTTPException(
            status_code=422,
            detail="Aucune action trouvée dans l'Excel (feuilles AA1–AA9 vides ou non reconnues). Les données existantes n'ont pas été modifiées.",
        )

    # Supprimer les anciennes actions et réinsérer
    db.query(models.ImprovementAction).filter(
        models.ImprovementAction.project_id == project_id
    ).delete(synchronize_session=False)

    now = datetime.now(timezone.utc).isoformat()
    for data in parsed:
        db.add(models.ImprovementAction(
            id=str(uuid4()),
            project_id=project_id,
            owner_id=current_user.id,
            created_at=now,
            **data,
        ))

    # Construire et stocker le résumé dans projects.excel_summary
    total_invest = sum(
        (d.get("investissement") or 0) for d in parsed if isinstance(d.get("investissement"), (int, float))
    )
    total_energie = sum(
        (d.get("economie_energie") or 0) for d in parsed if isinstance(d.get("economie_energie"), (int, float))
    )
    total_co2 = sum(
        (d.get("economie_co2") or 0) for d in parsed if isinstance(d.get("economie_co2"), (int, float))
    )
    excel_summary = {
        "imported_at": now,
        "nb_actions": len(parsed),
        "total_investissement_k_eur": round(total_invest, 2),
        "total_economie_energie_mwh_an": round(total_energie, 2),
        "total_economie_co2_kg_an": round(total_co2, 2),
        "actions": [
            {
                "reference": d.get("reference"),
                "intitule": d.get("intitule"),
                "classification": d.get("classification"),
                "type_amelioration": d.get("type_amelioration"),
                "investissement": d.get("investissement"),
                "economie_energie": d.get("economie_energie"),
                "economie_co2": d.get("economie_co2"),
            }
            for d in parsed
        ],
    }
    project.excel_summary = excel_summary
    flag_modified(project, "excel_summary")

    # Historique
    db.add(models.PlanAmeliorationHistory(
        id=str(uuid4()),
        project_id=project_id,
        owner_id=current_user.id,
        action_type="MANUAL_UPLOAD",
        changes={"excel_summary": excel_summary},
        created_at=now,
    ))

    db.commit()
    return {"imported": len(parsed), "actions": [d["intitule"] for d in parsed]}


# ══════════════════════════════════════════════════════════════════════════════
# ROUTES: Plan d'amélioration — import / preview  (AmurebaMappingService)
# ══════════════════════════════════════════════════════════════════════════════

_ALL_SEMANTIC_FIELDS = [
    "titre_action", "cout_investissement", "economie_energie_kwh",
    "economie_energie_euro", "taux_reduction_co2", "temps_retour_brut", "priorite",
]


async def _parse_uploaded_workbook(file: UploadFile):
    """
    Shared helper: read the uploaded .xlsx file and map all sheets.

    Returns (wb, mapped_sheets, filename) where mapped_sheets is the
    dict returned by AmurebaMappingService.map_workbook().

    Raises HTTPException 400 if the file is not a valid xlsx.
    """
    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True, keep_links=False)
    except Exception as exc:
        raise HTTPException(
            status_code=400,
            detail=f"Fichier xlsx invalide ou corrompu : {exc}",
        )
    svc = AmurebaMappingService()
    mapped = svc.map_workbook(wb)
    return wb, mapped, file.filename or "fichier.xlsx"


@app.get(
    "/projects/{project_id}/plan-amelioration",
    response_model=list[schemas.AmeliorationActionOut],
    summary="Lister les actions d'amélioration importées (JSONB)",
)
def list_plan_amelioration(
    project_id: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    project = db.query(models.Project).filter(
        models.Project.id == project_id,
        models.Project.owner_id == current_user.id,
    ).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    return db.query(models.AmeliorationAction).filter(
        models.AmeliorationAction.project_id == project_id
    ).order_by(models.AmeliorationAction.sheet_name).all()


@app.delete(
    "/projects/{project_id}/plan-amelioration/{action_id}",
    status_code=204,
    summary="Supprimer une action d'amélioration importée",
)
def delete_amelioration_action(
    project_id: str,
    action_id: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    project = db.query(models.Project).filter(
        models.Project.id == project_id,
        models.Project.owner_id == current_user.id,
    ).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    action = db.query(models.AmeliorationAction).filter(
        models.AmeliorationAction.id == action_id,
        models.AmeliorationAction.project_id == project_id,
    ).first()
    if not action:
        raise HTTPException(status_code=404, detail="Action not found")

    db.delete(action)
    db.commit()


@app.post(
    "/projects/{project_id}/plan-amelioration/preview",
    response_model=schemas.ImportPreviewResponse,
    summary="Prévisualiser un import AMUREBA sans sauvegarder",
)
async def preview_plan_amelioration(
    project_id: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """
    Read an AMUREBA xlsx and return a structured preview — nothing is saved.

    Use this endpoint to let the frontend show the user which sheets were
    detected, which semantic fields were mapped, and what data will be
    imported before asking for confirmation.

    Returns:
        ImportPreviewResponse with per-sheet summaries.
    """
    project = db.query(models.Project).filter(
        models.Project.id == project_id,
        models.Project.owner_id == current_user.id,
    ).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    _, mapped, filename = await _parse_uploaded_workbook(file)

    sheets_out = []
    total_rows = 0
    for sheet_name, sheet_data in mapped.items():
        kv = sheet_data.get("key_values", {})
        rows = sheet_data.get("raw_rows", [])
        headers = sheet_data.get("detected_headers", [])
        unmapped = sheet_data.get("unmapped_headers", [])
        missing = [f for f in _ALL_SEMANTIC_FIELDS if f not in kv]
        total_rows += len(rows)
        sheets_out.append(schemas.ImportPreviewSheet(
            sheet_name=sheet_name,
            row_count=len(rows),
            detected_headers=headers,
            key_values=kv,
            unmapped_headers=unmapped,
            missing_semantic_fields=missing,
        ))

    return schemas.ImportPreviewResponse(
        filename=filename,
        sheet_names=list(mapped.keys()),
        sheets=sheets_out,
        total_rows=total_rows,
    )


@app.post(
    "/projects/{project_id}/plan-amelioration/import",
    response_model=schemas.ImportResponse,
    status_code=201,
    summary="Importer un Excel AMUREBA et sauvegarder en base",
)
async def import_plan_amelioration(
    project_id: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """
    Import an AMUREBA xlsx: parse all sheets, map semantic fields,
    then **replace** all existing amelioration_actions for this project.

    Workflow:
        1. Validate xlsx (400 if invalid)
        2. Map all non-empty sheets via AmurebaMappingService
        3. Delete previous amelioration_actions for this project
        4. Insert one AmeliorationAction row per mapped sheet
        5. Return import summary

    The full AmurebaMappingService result is stored as-is in action_data
    (JSONB), so no information is lost and the schema can evolve without
    a migration.
    """
    project = db.query(models.Project).filter(
        models.Project.id == project_id,
        models.Project.owner_id == current_user.id,
    ).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    _, mapped, filename = await _parse_uploaded_workbook(file)

    if not mapped:
        raise HTTPException(
            status_code=422,
            detail="Aucune feuille avec des données n'a été trouvée dans ce fichier.",
        )

    now = datetime.now(timezone.utc).isoformat()

    # Replace previous import
    db.query(models.AmeliorationAction).filter(
        models.AmeliorationAction.project_id == project_id
    ).delete(synchronize_session=False)

    for sheet_name, sheet_data in mapped.items():
        db.add(models.AmeliorationAction(
            id=str(uuid4()),
            project_id=project_id,
            sheet_name=sheet_name,
            action_data=sheet_data,
            created_at=now,
            updated_at=now,
        ))

    db.commit()

    return schemas.ImportResponse(
        imported_sheets=len(mapped),
        sheet_names=list(mapped.keys()),
        filename=filename,
        imported_at=now,
    )



# ROUTES: LCA
# ==============================
@app.get("/lca/materials", response_model=List[schemas.LcaMaterialOut])
def list_lca_materials(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    return db.query(models.LcaMaterial).order_by(models.LcaMaterial.category, models.LcaMaterial.name).all()


@app.get("/projects/{project_id}/lca")
def get_project_lca(
    project_id: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    project = db.query(models.Project).filter(
        models.Project.id == project_id,
        models.Project.owner_id == current_user.id,
    ).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    lca = db.query(models.LcaProject).filter(models.LcaProject.project_id == project_id).first()

    # Fetch optimisation cache columns (added by migration 010) via raw SQL
    # so we don't need to modify models.py
    opt_row = db.execute(
        _sa_text(
            "SELECT optimisation_hash, optimisation_cache "
            "FROM lca_projects WHERE project_id = :pid"
        ),
        {"pid": project_id},
    ).fetchone()
    opt_hash  = opt_row[0] if opt_row else None
    opt_cache = opt_row[1] if opt_row else None

    return {
        "batiments": lca.batiments if lca else [],
        # legacy fields kept for backward compatibility
        "elements": lca.elements if lca else [],
        "parois":   lca.parois   if lca else [],
        "batiment": lca.batiment if lca else {},
        # optimisation cache
        "optimisation_hash":  opt_hash,
        "optimisation_cache": opt_cache,
    }


@app.patch("/projects/{project_id}/lca")
def update_project_lca(
    project_id: str,
    payload: schemas.LcaProjectUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """Legacy route — sauvegarde uniquement les éléments ACV (ancien format)."""
    project = db.query(models.Project).filter(
        models.Project.id == project_id,
        models.Project.owner_id == current_user.id,
    ).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    now = datetime.now(timezone.utc).isoformat()
    elements = [e.model_dump() for e in payload.elements]

    lca = db.query(models.LcaProject).filter(models.LcaProject.project_id == project_id).first()
    if lca:
        lca.elements = elements
        lca.updated_at = now
        flag_modified(lca, "elements")
    else:
        lca = models.LcaProject(
            id=str(uuid4()),
            project_id=project_id,
            elements=elements,
            created_at=now,
            updated_at=now,
        )
        db.add(lca)

    db.commit()
    db.refresh(lca)
    return {"elements": lca.elements}


class _LcaBatimentsPayload(_PydanticBase):
    batiments: List[Dict[str, Any]]


@app.patch("/projects/{project_id}/lca/batiments")
def update_project_lca_batiments(
    project_id: str,
    payload: _LcaBatimentsPayload,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """Sauvegarde le tableau complet des bâtiments avec parois, composants et paramètres."""
    project = db.query(models.Project).filter(
        models.Project.id == project_id,
        models.Project.owner_id == current_user.id,
    ).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    now = datetime.now(timezone.utc).isoformat()
    lca = db.query(models.LcaProject).filter(models.LcaProject.project_id == project_id).first()
    if lca:
        lca.batiments = payload.batiments
        lca.updated_at = now
        flag_modified(lca, "batiments")
    else:
        lca = models.LcaProject(
            id=str(uuid4()),
            project_id=project_id,
            elements=[],
            parois=[],
            batiment={},
            batiments=payload.batiments,
            created_at=now,
            updated_at=now,
        )
        db.add(lca)

    db.commit()
    db.refresh(lca)
    return {"batiments": lca.batiments}


class _LcaOptCachePayload(_PydanticBase):
    hash: str
    cache: Dict[str, Any]


@app.patch("/projects/{project_id}/lca/optimisation-cache")
def patch_lca_optimisation_cache(
    project_id: str,
    payload: _LcaOptCachePayload,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """Sauvegarde le hash de configuration et les 5 solutions phares après optimisation."""
    project = db.query(models.Project).filter(
        models.Project.id == project_id,
        models.Project.owner_id == current_user.id,
    ).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    now = datetime.now(timezone.utc).isoformat()

    existing = db.execute(
        _sa_text("SELECT id FROM lca_projects WHERE project_id = :pid"),
        {"pid": project_id},
    ).fetchone()

    if existing:
        db.execute(
            _sa_text(
                "UPDATE lca_projects "
                "SET optimisation_hash = :h, optimisation_cache = :c::jsonb, updated_at = :now "
                "WHERE project_id = :pid"
            ),
            {"h": payload.hash, "c": json.dumps(payload.cache), "now": now, "pid": project_id},
        )
    else:
        # No LCA row yet — create a minimal one so the cache isn't lost
        db.execute(
            _sa_text(
                "INSERT INTO lca_projects "
                "(id, project_id, elements, parois, batiment, batiments, "
                " optimisation_hash, optimisation_cache, created_at, updated_at) "
                "VALUES (:id, :pid, '[]'::jsonb, '[]'::jsonb, '{}'::jsonb, '[]'::jsonb, "
                "        :h, :c::jsonb, :now, :now)"
            ),
            {
                "id": str(uuid4()), "pid": project_id,
                "h": payload.hash, "c": json.dumps(payload.cache),
                "now": now,
            },
        )

    db.commit()
    return {"status": "ok"}


# ──────────────────────────────────────────────────────────────────────────────
# HELPER : parseur LCIA-results.xlsx → dict d'impacts EF v3.0
# ──────────────────────────────────────────────────────────────────────────────

# Patterns ordonnés du plus spécifique au plus général.
# On cherche chaque sous-chaîne dans le nom de colonne normalisé (lowercase, espaces normalisés).
# Les sous-catégories (biogenic, fossil…) doivent précéder leur catégorie parente (climate change).
_EF_COLUMN_PATTERNS: list = [
    # Climate change — sous-catégories d'abord
    ("climate change: biogenic",                        "climate_biogenic"),
    ("climate change: fossil",                          "climate_fossil"),
    ("climate change: land use",                        "climate_landuse"),
    # Human toxicity — carcinogène avant non-carcinogène (évite faux positif sur "non")
    ("human toxicity: carcinogenic",                    "human_tox_carc"),
    ("human toxicity: non-carcinogenic",                "human_tox_noncarc"),
    # Ecotoxicity freshwater
    ("ecotoxicity: freshwater",                         "ecotoxicity_fw"),
    # Eutrophication
    ("eutrophication: freshwater",                      "eutrophication_fw"),
    ("eutrophication: marine",                          "eutrophication_marine"),
    ("eutrophication: terrestrial",                     "eutrophication_terrestrial"),
    # Autres indicateurs spécifiques
    ("ionising radiation",                              "ionising_radiation"),
    ("ionizing radiation",                              "ionising_radiation"),
    ("photochemical oxidant",                           "photochemical_oxidant"),
    ("particulate matter",                              "particulate_matter"),
    ("ozone depletion",                                 "ozone_depletion"),
    ("energy resources: non-renewable",                 "energy_nonrenewable"),
    ("material resources: metals/minerals",             "material_resources"),
    ("material resources",                              "material_resources"),
    ("| land use |",                                    "land_use"),   # pipes pour éviter collision avec "climate change: land use"
    ("water use",                                       "water_use"),
    ("acidification",                                   "acidification"),
    # Climate change général (GWP100) — en dernier pour ne pas écraser les sous-catégories
    ("climate change",                                  "gwp100"),
]


def _normalize_col(name: str) -> str:
    """Normalise un nom de colonne : lowercase + espaces multiples → un seul espace."""
    import re as _re
    return _re.sub(r"\s+", " ", name.lower().strip())


def _parse_lcia_xlsx(file_bytes: bytes) -> Dict[str, float]:
    """Lit un fichier LCIA-results.xlsx (colonnes EF v3.0) et retourne {clé_ef: valeur}."""
    import math
    import io
    import pandas as pd

    impacts: Dict[str, float] = {}
    xl = pd.ExcelFile(io.BytesIO(file_bytes))

    for sheet_name in xl.sheet_names:
        # Essaie de trouver la ligne d'en-tête qui contient "EF v3.0"
        df = None
        for header_row in range(6):
            candidate = xl.parse(sheet_name, header=header_row)
            ef_cols = [
                c for c in candidate.columns
                if isinstance(c, str) and "EF v3.0" in c
            ]
            if ef_cols:
                df = candidate
                break

        if df is None or df.empty:
            continue

        ef_cols = [c for c in df.columns if isinstance(c, str) and "EF v3.0" in c]

        for col in ef_cols:
            col_norm = _normalize_col(col)
            key = None
            for substr, k in _EF_COLUMN_PATTERNS:
                if substr in col_norm:
                    key = k
                    break
            if not key:
                continue

            # Première valeur numérique valide dans la colonne
            for val in df[col]:
                if isinstance(val, (int, float)) and not (
                    isinstance(val, float) and math.isnan(val)
                ):
                    impacts[key] = float(val)
                    break

    return impacts


# ==============================
# ROUTES: LCA ADMIN
# ==============================

@app.post("/lca/materials/import", response_model=schemas.LcaMaterialOut)
async def import_lca_material(
    file: UploadFile = File(...),
    name: str = Form(...),
    category: str = Form(...),
    functional_unit: str = Form(...),
    unit: str = Form(...),
    prix: float = Form(...),
    valeur_r: float = Form(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Seuls les fichiers .xlsx sont acceptés.")

    file_bytes = await file.read()

    try:
        impacts = _parse_lcia_xlsx(file_bytes)
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"Erreur de lecture du fichier : {str(e)}")

    if not impacts:
        raise HTTPException(
            status_code=422,
            detail="Aucun indicateur EF v3.0 reconnu dans ce fichier. Vérifiez que les noms de catégories d'impact correspondent au standard EF v3.0.",
        )

    material = models.LcaMaterial(
        id=str(uuid4()),
        name=name,
        category=category,
        functional_unit=functional_unit,
        unit=unit,
        impacts=impacts,
        prix=prix,
        valeur_r=valeur_r,
    )
    db.add(material)
    db.commit()
    db.refresh(material)
    return material


@app.get("/lca/materials/{material_id}", response_model=schemas.LcaMaterialOut)
def get_lca_material(
    material_id: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    material = db.query(models.LcaMaterial).filter(models.LcaMaterial.id == material_id).first()
    if not material:
        raise HTTPException(status_code=404, detail="Matériau introuvable")
    return material


@app.patch("/lca/materials/{material_id}", response_model=schemas.LcaMaterialOut)
def patch_lca_material(
    material_id: str,
    payload: _LcaMaterialEditPayload,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    material = db.query(models.LcaMaterial).filter(models.LcaMaterial.id == material_id).first()
    if not material:
        raise HTTPException(status_code=404, detail="Matériau introuvable")
    for field, value in payload.model_dump(exclude_none=True).items():
        setattr(material, field, value)
    db.commit()
    db.refresh(material)
    return material


@app.delete("/lca/materials/{material_id}")
def delete_lca_material(
    material_id: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    material = db.query(models.LcaMaterial).filter(models.LcaMaterial.id == material_id).first()
    if not material:
        raise HTTPException(status_code=404, detail="Matériau introuvable")
    db.delete(material)
    db.commit()
    return {"status": "deleted", "id": material_id}


@app.post("/lca/materials/{material_id}/duplicate", response_model=schemas.LcaMaterialOut)
def duplicate_lca_material(
    material_id: str,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    original = db.query(models.LcaMaterial).filter(models.LcaMaterial.id == material_id).first()
    if not original:
        raise HTTPException(status_code=404, detail="Matériau introuvable")
    copy = models.LcaMaterial(
        id=str(uuid4()),
        name=original.name + " (copie)",
        category=original.category,
        functional_unit=original.functional_unit,
        unit=original.unit,
        impacts=dict(original.impacts),
        prix=original.prix,
        valeur_r=original.valeur_r,
        is_fixed=False,
        flux_reference=original.flux_reference,
    )
    db.add(copy)
    db.commit()
    db.refresh(copy)
    return copy
