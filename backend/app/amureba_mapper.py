"""
AmurebaMappingService
=====================
Maps raw openpyxl worksheet data to semantic improvement-action fields.

Two complementary strategies are applied per sheet:
  1. Fixed-cell extraction  — reads cells at known AMUREBA template positions
     (B9, G61, G77, G87, N15, …).  Reliable for well-formed AA1-AA9 sheets.
  2. Tabular extraction     — auto-detects the header row, extracts data rows,
     then fuzzy-maps each header to a semantic field.  Works even when the
     auditor renames or moves columns.

Fuzzy matching uses difflib.SequenceMatcher (stdlib, no extra dependency).
To switch to rapidfuzz for faster performance on large files:
    pip install rapidfuzz
    # Then replace fuzzy_score() body with:
    #   from rapidfuzz import fuzz
    #   return fuzz.token_set_ratio(normalize_str(a), normalize_str(b)) / 100

Adding aliases:
    Extend SEMANTIC_ALIASES with new entries, or pass extra_aliases=
    to AmurebaMappingService.__init__().

Edge cases guaranteed never to crash:
  - Empty sheet           → map_sheet() returns None
  - Excel error strings   → safe_cell_value() returns None
  - Missing cells         → caught per-cell; rest of sheet still processed
  - Sheets absent from workbook → map_workbook() logs warning, continues
"""

from __future__ import annotations

import logging
import re
import unicodedata
from difflib import SequenceMatcher
from typing import Any, Dict, List, Optional, Tuple

from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

# ──────────────────────────────────────────────────────────────────────────────
# Constants
# ──────────────────────────────────────────────────────────────────────────────

#: Excel formula-error prefixes we treat as "no value"
EXCEL_ERROR_PREFIXES = frozenset(
    {"#REF!", "#VALUE!", "#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!", "#GETTING_DATA"}
)

#: Minimum SequenceMatcher ratio to consider a header-to-alias match valid
FUZZY_THRESHOLD = 0.65

#: Fixed cell addresses in AMUREBA AA1-AA9 form sheets → internal semantic key.
#: Values marked with "_raw_*" are post-processed (see extract_amureba_cells).
AMUREBA_FORM_CELLS: Dict[str, str] = {
    "F5":  "reference",             # N° amélioration (AA1, AA2…)
    "B9":  "titre_action",          # Intitulé de l'action
    "B13": "type_amelioration",     # Type (dropdown value)
    "F27": "classification",        # A ou B
    "G61": "cout_investissement",   # Total investissement €
    "G77": "_mwh_an",               # Économie énergie finale MWh/an  →  ×1000 = kWh
    "G87": "taux_reduction_co2",    # Réductions GES  kg CO2/an
    "N10": "_irr_avant",            # IRR avant impôt (num or "PROJET NON RENTABLE")
    "N15": "temps_retour_brut",     # PBT avant impôt  (années)
    "N17": "_irr_apres",            # IRR après impôt
    "N18": "_pbt_apres",            # PBT après impôt
    "K18": "duree_amortissement",   # Durée d'amortissement estimée (années)
    "K22": "_ets_raw",              # Entreprise ETS ? (OUI/NON)
    "K23": "_deduction_raw",        # Déduction fiscale ? (OUI/NON)
}

#: Semantic fields exposed to callers and their known column-name aliases.
#: Edit freely — the fuzzy matcher handles partial or inexact matches.
SEMANTIC_ALIASES: Dict[str, List[str]] = {
    "titre_action": [
        "titre", "intitulé", "intitule", "libellé", "libelle",
        "action", "objet", "nom de l action", "désignation", "description",
    ],
    "cout_investissement": [
        "coût", "cout", "investissement", "invest", "capex",
        "coût investissement", "coût total", "montant", "invest €",
        "capital", "dépense", "depense",
    ],
    "economie_energie_kwh": [
        "économie énergie kwh", "economie energie kwh", "économie kwh",
        "gain énergie kwh", "gain kwh", "économie annuelle kwh",
        "réduction consommation", "ep kwh", "ef kwh",
        "économie énergie", "economie energie", "gain energie",
    ],
    "economie_energie_euro": [
        "économie euro", "economie euro", "gain financier",
        "économie annuelle", "économie facture", "gain €",
        "économie en €", "gain en euro", "économie financière",
    ],
    "taux_reduction_co2": [
        "co2", "réduction co2", "reduction co2", "émissions",
        "taux co2", "kg co2", "kg co2 an", "co2 an",
        "ges", "réduction ges", "reduction ges", "émissions ges",
    ],
    "temps_retour_brut": [
        "pbt", "payback", "temps retour", "retour investissement",
        "trs", "délai", "délai retour", "trb", "temps de retour",
        "retour simple", "ann retour", "durée retour",
    ],
    "priorite": [
        "priorité", "priorite", "urgence", "rang", "ordre",
        "classification", "niveau", "priorité action",
    ],
}


# ──────────────────────────────────────────────────────────────────────────────
# String utilities
# ──────────────────────────────────────────────────────────────────────────────

def normalize_str(s: Any) -> str:
    """
    Prepare a string for fuzzy comparison.

    Steps:
      1. Cast to str and strip
      2. Decompose accented chars and drop combining marks  (é→e, ê→e …)
      3. Lowercase
      4. Replace every non-alphanumeric character with a space
      5. Collapse consecutive spaces

    Examples:
        "Coût €"                  →  "cout"
        "économie d'énergie kWh"  →  "economie d energie kwh"
        "  N/A  "                 →  "n a"
        42                        →  "42"
    """
    if s is None:
        return ""
    nfkd = unicodedata.normalize("NFKD", str(s).strip())
    ascii_only = nfkd.encode("ascii", "ignore").decode("ascii")
    cleaned = re.sub(r"[^a-z0-9]", " ", ascii_only.lower())
    return re.sub(r"\s+", " ", cleaned).strip()


def fuzzy_score(a: Any, b: Any) -> float:
    """
    Return a similarity ratio in [0.0, 1.0] between two values after
    normalizing both with normalize_str().

    Uses difflib.SequenceMatcher (no external dependency).

    Edge cases:
        ""  vs anything  →  0.0
        "x" vs "x"       →  1.0
    """
    na, nb = normalize_str(a), normalize_str(b)
    if not na or not nb:
        return 0.0
    return SequenceMatcher(None, na, nb).ratio()


def _best_match_for_header(
    header: str,
    field_aliases: Dict[str, List[str]],
) -> Tuple[Optional[str], float]:
    """
    Find the semantic field whose aliases best match *header*.

    Returns (field_name, score) or (None, 0.0) if nothing exceeds
    FUZZY_THRESHOLD.

    Algorithm per alias:
      - Exact normalized match  → score = 1.0  (early return)
      - One is a substring of the other → boost proportional to length overlap
      - Otherwise SequenceMatcher ratio
    """
    best_field: Optional[str] = None
    best_score: float = 0.0
    nh = normalize_str(header)

    for field, aliases in field_aliases.items():
        for alias in aliases:
            na = normalize_str(alias)
            if not na:
                continue
            if nh == na:
                return field, 1.0
            if na in nh or nh in na:
                overlap = min(len(na), len(nh)) / max(len(na), len(nh))
                score = max(overlap, SequenceMatcher(None, nh, na).ratio())
            else:
                score = SequenceMatcher(None, nh, na).ratio()
            if score > best_score:
                best_score = score
                best_field = field

    if best_score >= FUZZY_THRESHOLD:
        return best_field, best_score
    return None, 0.0


# ──────────────────────────────────────────────────────────────────────────────
# Cell-level helpers
# ──────────────────────────────────────────────────────────────────────────────

def safe_cell_value(cell) -> Any:
    """
    Safely read an openpyxl cell value.

    Returns None for:
      - Empty cells
      - Excel error strings (#REF!, #VALUE!, #DIV/0! …)
      - Whitespace-only strings

    Never raises — all exceptions are caught and logged.
    """
    try:
        v = cell.value
    except Exception as exc:  # noqa: BLE001
        log.debug("Could not read cell %s: %s", getattr(cell, "coordinate", "?"), exc)
        return None

    if v is None:
        return None
    if isinstance(v, str):
        stripped = v.strip()
        if not stripped:
            return None
        if stripped.startswith("#") or stripped in EXCEL_ERROR_PREFIXES:
            log.debug("Excel error at %s: %r — replaced with null",
                      getattr(cell, "coordinate", "?"), stripped)
            return None
        return stripped
    return v


def detect_cell_type(v: Any) -> str:
    """Return a type hint string: 'number', 'text', 'date', 'bool', or 'null'."""
    if v is None:
        return "null"
    if isinstance(v, bool):
        return "bool"
    if isinstance(v, (int, float)):
        return "number"
    type_name = type(v).__name__.lower()
    if "date" in type_name or "time" in type_name:
        return "date"
    return "text"


# ──────────────────────────────────────────────────────────────────────────────
# Tabular-region detection
# ──────────────────────────────────────────────────────────────────────────────

def _string_ratio(values: List[Any]) -> float:
    """Fraction of non-None values that are strings."""
    non_empty = [v for v in values if v is not None]
    if not non_empty:
        return 0.0
    return sum(1 for v in non_empty if isinstance(v, str)) / len(non_empty)


def detect_header_row(ws, max_scan_rows: int = 25) -> Optional[int]:
    """
    Heuristically identify the header row in a worksheet.

    A row qualifies as "header" when:
      - It has ≥ 2 non-empty cells, AND
      - ≥ 60 % of those cells are strings (not numbers/dates)

    Scans at most *max_scan_rows* from the top.

    Returns:
        1-based row index of the header row, or None.

    Edge cases:
        - Entirely numeric first rows → scans deeper
        - Merged cells → openpyxl returns None for merged slaves; ignored
        - max_column = 0 → returns None
    """
    max_col = ws.max_column or 0
    if max_col == 0:
        return None

    for row_idx in range(1, min(max_scan_rows, ws.max_row or 0) + 1):
        row_vals = [
            safe_cell_value(ws.cell(row=row_idx, column=c))
            for c in range(1, max_col + 1)
        ]
        non_empty = [v for v in row_vals if v is not None]
        if len(non_empty) >= 2 and _string_ratio(row_vals) >= 0.60:
            return row_idx
    return None


def extract_tabular_data(ws) -> Dict[str, Any]:
    """
    Extract tabular data from a worksheet.

    1. Auto-detects the header row (or falls back to column letters).
    2. Reads every subsequent non-empty row as a dict {header: value}.
    3. Skips rows where all values are None.

    Returns:
        {
          "detected_headers":  [str, …],
          "raw_rows":          [{"row_index": int, "values": {header: value}}, …],
          "header_row_index":  int | None,
        }

    Edge cases:
        - Empty rows interspersed → silently skipped
        - Header cell empty → column letter used as fallback key
        - max_column / max_row = 0 or None → returns empty result
    """
    max_col = ws.max_column or 0
    max_row = ws.max_row or 0

    if max_col == 0 or max_row == 0:
        return {"detected_headers": [], "raw_rows": [], "header_row_index": None}

    header_row_idx = detect_header_row(ws)

    # Build headers list; track which entries are genuine values vs. fallbacks.
    # is_actual[i] = True  → header came from a real cell value
    # is_actual[i] = False → it is a column-letter placeholder (e.g. "A", "BC")
    if header_row_idx is None:
        # No header detected — use column letters as internal keys only
        headers = [get_column_letter(c) for c in range(1, max_col + 1)]
        is_actual: List[bool] = [False] * max_col
        data_start = 1
    else:
        headers = []
        is_actual = []
        for c in range(1, max_col + 1):
            v = safe_cell_value(ws.cell(row=header_row_idx, column=c))
            if v is not None:
                headers.append(str(v))
                is_actual.append(True)
            else:
                headers.append(get_column_letter(c))
                is_actual.append(False)
        data_start = header_row_idx + 1

    raw_rows: List[Dict[str, Any]] = []
    for row_idx in range(data_start, max_row + 1):
        row_dict: Dict[str, Any] = {}
        has_value = False
        for col_idx in range(1, max_col + 1):
            v = safe_cell_value(ws.cell(row=row_idx, column=col_idx))
            key = headers[col_idx - 1] if col_idx - 1 < len(headers) else get_column_letter(col_idx)
            row_dict[key] = v
            if v is not None:
                has_value = True
        if has_value:
            raw_rows.append({"row_index": row_idx, "values": row_dict})

    # Only expose headers that came from real cell content (not fallbacks).
    detected = [h for h, actual in zip(headers, is_actual) if actual]

    return {
        "detected_headers": detected,
        "raw_rows": raw_rows,
        "header_row_index": header_row_idx,
    }


# ──────────────────────────────────────────────────────────────────────────────
# AmurebaMappingService
# ──────────────────────────────────────────────────────────────────────────────

class AmurebaMappingService:
    """
    Maps raw Excel workbook data to semantic improvement-action fields.

    Instantiate once; call map_workbook() or map_sheet() as needed.

    Args:
        extra_aliases: optional {semantic_field: [alias, …]} merged with
                       the built-in SEMANTIC_ALIASES.  Use to add project-
                       specific column names without editing the module.

    Example::

        svc = AmurebaMappingService(
            extra_aliases={"cout_investissement": ["budget total"]}
        )
        with open("AMUREBA.xlsx", "rb") as f:
            wb = load_workbook(f, data_only=True)
        result = svc.map_workbook(wb)
    """

    def __init__(self, extra_aliases: Optional[Dict[str, List[str]]] = None):
        self._aliases: Dict[str, List[str]] = {
            field: list(vals) for field, vals in SEMANTIC_ALIASES.items()
        }
        if extra_aliases:
            for field, alts in extra_aliases.items():
                self._aliases.setdefault(field, []).extend(alts)

    # ── Public API ─────────────────────────────────────────────────────────────

    def map_headers(self, headers: List[str]) -> Dict[str, str]:
        """
        Map raw column headers to semantic field names using fuzzy matching.

        Returns:
            {raw_header: semantic_field} for every matched header.
            Unmatched headers are omitted (check unmapped_headers in map_sheet).

        Example:
            ["Coût total", "économie énergie (kWh)", "Délai retour"]
            → {"Coût total": "cout_investissement",
               "économie énergie (kWh)": "economie_energie_kwh",
               "Délai retour": "temps_retour_brut"}

        Edge cases:
            - Duplicate headers → last-seen wins
            - None / empty string headers → skipped
            - Score below FUZZY_THRESHOLD → not included
        """
        result: Dict[str, str] = {}
        for header in headers:
            if not header:
                continue
            field, _ = _best_match_for_header(header, self._aliases)
            if field:
                result[str(header)] = field
        return result

    def extract_amureba_cells(self, ws) -> Dict[str, Any]:
        """
        Read key values from fixed AMUREBA template cell positions.

        Post-processing applied automatically:
          - _mwh_an (G77)      → economie_energie_kwh = value × 1 000
          - _irr_avant (N10)   → irr_avant_impot (None if text)
          - _irr_apres (N17)   → irr_apres_impot (None if text)
          - _pbt_apres (N18)   → pbt_apres_impot
          - _ets_raw (K22)     → entreprise_ets  True/False/None
          - _deduction_raw     → deduction_fiscale  True/False/None

        Returns:
            {semantic_field: value} — only fields with non-None values.

        Edge cases:
            - Cells missing from sheet → silently skipped
            - IRR cell = "PROJET NON RENTABLE" → stored as None
            - OUI/NON boolean parsing is case-insensitive
        """
        raw: Dict[str, Any] = {}
        for addr, key in AMUREBA_FORM_CELLS.items():
            try:
                raw[key] = safe_cell_value(ws[addr])
            except Exception:  # noqa: BLE001
                pass

        result: Dict[str, Any] = {}

        # Direct fields
        for direct_key in (
            "reference", "titre_action", "type_amelioration", "classification",
            "cout_investissement", "taux_reduction_co2", "temps_retour_brut",
            "duree_amortissement",
        ):
            v = raw.get(direct_key)
            if v is not None:
                result[direct_key] = v

        # MWh → kWh conversion
        mwh = raw.get("_mwh_an")
        if isinstance(mwh, (int, float)):
            result["economie_energie_kwh"] = mwh * 1_000
            result["_economie_energie_mwh"] = mwh   # keep raw for reference

        # IRR fields: numeric or None (text values like "PROJET NON RENTABLE" → None)
        for src_key, dest_key in (
            ("_irr_avant", "irr_avant_impot"),
            ("_irr_apres", "irr_apres_impot"),
            ("_pbt_apres", "pbt_apres_impot"),
        ):
            v = raw.get(src_key)
            result[dest_key] = v if isinstance(v, (int, float)) else None

        # OUI/NON booleans
        for src_key, dest_key in (
            ("_ets_raw", "entreprise_ets"),
            ("_deduction_raw", "deduction_fiscale"),
        ):
            v = raw.get(src_key)
            if v is None:
                result[dest_key] = None
            else:
                result[dest_key] = str(v).strip().upper() == "OUI"

        return {k: v for k, v in result.items() if v is not None}

    def map_sheet(self, ws, sheet_name: str) -> Optional[Dict[str, Any]]:
        """
        Full mapping pipeline for one worksheet.

        Steps:
          1. Emptiness check — returns None if sheet has no values
          2. Tabular extraction (header detection + data rows)
          3. Fixed-cell AMUREBA extraction
          4. Fuzzy header mapping
          5. Merge: tabular key_values ← AMUREBA cells (fixed cells win)

        Returns:
            None if the sheet is empty, otherwise::

                {
                    "raw_rows":        [{"row_index": int, "values": {…}}, …],
                    "detected_headers": [str, …],
                    "header_mapping":  {raw_header: semantic_field},
                    "key_values":      {semantic_field: value},
                    "unmapped_headers": [str, …],   # headers with no fuzzy match
                }

        Edge cases:
            - Sheet with only headers, no data → raw_rows = []
            - AMUREBA cells take precedence over tabular extraction
            - All exceptions inside are caught per-cell; the sheet is never skipped
              due to a single bad cell
        """
        max_row = ws.max_row or 0
        if max_row == 0:
            log.info("Sheet %r: max_row=0, skipping", sheet_name)
            return None

        # Quick check: does the sheet have any non-None value?
        has_any = any(
            safe_cell_value(cell) is not None
            for row in ws.iter_rows()
            for cell in row
        )
        if not has_any:
            log.info("Sheet %r: all cells empty, skipping", sheet_name)
            return None

        # 1. Tabular extraction
        tabular = extract_tabular_data(ws)
        log.info(
            "Sheet %r: %d header(s) detected, %d data row(s)",
            sheet_name, len(tabular["detected_headers"]), len(tabular["raw_rows"]),
        )

        # 2. Fuzzy header mapping
        header_mapping = self.map_headers(tabular["detected_headers"])
        unmapped = [h for h in tabular["detected_headers"] if h not in header_mapping]

        # 3. Build key_values from first data row + header mapping
        key_values: Dict[str, Any] = {}
        if tabular["raw_rows"] and header_mapping:
            first_row = tabular["raw_rows"][0]["values"]
            for raw_hdr, sem_field in header_mapping.items():
                v = first_row.get(raw_hdr)
                if v is not None:
                    key_values[sem_field] = v

        # 4. AMUREBA fixed-cell extraction (overwrites tabular where both exist)
        amureba_kv = self.extract_amureba_cells(ws)
        key_values.update(amureba_kv)

        return {
            "raw_rows": tabular["raw_rows"],
            "detected_headers": tabular["detected_headers"],
            "header_mapping": header_mapping,
            "key_values": key_values,
            "unmapped_headers": unmapped,
        }

    def map_workbook(self, wb) -> Dict[str, Any]:
        """
        Map all non-empty sheets of a workbook.

        Returns:
            {sheet_name: map_sheet_result} — only sheets with content.

        Edge cases:
            - A sheet that raises an unexpected exception → warning logged,
              sheet skipped, other sheets still processed
            - Empty sheets → silently omitted
        """
        results: Dict[str, Any] = {}
        for sheet_name in wb.sheetnames:
            try:
                ws = wb[sheet_name]
                sheet_result = self.map_sheet(ws, sheet_name)
                if sheet_result is not None:
                    results[sheet_name] = sheet_result
            except Exception as exc:  # noqa: BLE001
                log.warning("Unexpected error processing sheet %r: %s", sheet_name, exc)
        return results
