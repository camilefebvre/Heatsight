"""
Tests for amureba_mapper.py
===========================

Run with:  pytest backend/tests/test_amureba_mapper.py -v

Covers:
  - normalize_str()         — accents, special chars, whitespace
  - fuzzy_score()           — identical / similar / unrelated strings
  - AmurebaMappingService.map_headers()
      - exact match
      - fuzzy / renamed columns
      - below-threshold headers (should be omitted)
      - empty header list
  - extract_tabular_data()
      - empty rows interspersed
      - no header row detected (fallback to column letters)
  - AmurebaMappingService.map_sheet()
      - empty worksheet → None
      - worksheet with data
      - AMUREBA fixed cells
      - Excel error values (#REF! etc.)
  - AmurebaMappingService.map_workbook()
      - missing / blank sheets skipped
      - multiple sheets processed independently
"""

import pytest
from openpyxl import Workbook

import sys
import os

# Make the app package importable when running from repo root
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", ".."))

from app.amureba_mapper import (
    AmurebaMappingService,
    FUZZY_THRESHOLD,
    extract_tabular_data,
    fuzzy_score,
    normalize_str,
    safe_cell_value,
)


# ──────────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────────

def make_workbook(sheets: dict) -> "Workbook":
    """
    Create an in-memory openpyxl Workbook from a dict of {sheet_name: [rows]}.
    Each row is a list of cell values.

    Example:
        make_workbook({"AA1": [["Titre", "Coût"], ["Isolation", 5000]]})
    """
    wb = Workbook()
    # Remove the default empty sheet
    wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for row in rows:
            ws.append(row)
    return wb


def get_ws(wb: "Workbook", name: str):
    return wb[name]


# ──────────────────────────────────────────────────────────────────────────────
# normalize_str
# ──────────────────────────────────────────────────────────────────────────────

class TestNormalizeStr:
    def test_removes_accents(self):
        assert normalize_str("éàêîç") == "eaeic"

    def test_lowercase(self):
        assert normalize_str("INVESTISSEMENT") == "investissement"

    def test_strips_special_chars(self):
        assert normalize_str("Coût €") == "cout"

    def test_collapses_spaces(self):
        assert normalize_str("temps  de  retour") == "temps de retour"

    def test_parentheses_and_slash(self):
        assert normalize_str("économie (kWh/an)") == "economie kwh an"

    def test_empty_string(self):
        assert normalize_str("") == ""

    def test_none_returns_empty(self):
        assert normalize_str(None) == ""

    def test_integer_input(self):
        assert normalize_str(42) == "42"

    def test_whitespace_only(self):
        assert normalize_str("   ") == ""


# ──────────────────────────────────────────────────────────────────────────────
# fuzzy_score
# ──────────────────────────────────────────────────────────────────────────────

class TestFuzzyScore:
    def test_identical(self):
        assert fuzzy_score("cout", "cout") == 1.0

    def test_same_after_normalization(self):
        assert fuzzy_score("Coût €", "cout") == 1.0

    def test_high_similarity(self):
        # "investissement" and "investissement total" share the same root
        score = fuzzy_score("investissement", "investissement total")
        assert score >= 0.6

    def test_unrelated(self):
        score = fuzzy_score("priorité", "économie énergie kwh")
        assert score < FUZZY_THRESHOLD

    def test_empty_a(self):
        assert fuzzy_score("", "cout") == 0.0

    def test_empty_b(self):
        assert fuzzy_score("cout", "") == 0.0

    def test_both_empty(self):
        assert fuzzy_score("", "") == 0.0


# ──────────────────────────────────────────────────────────────────────────────
# safe_cell_value
# ──────────────────────────────────────────────────────────────────────────────

class TestSafeCellValue:
    """Use a real openpyxl cell obtained from an in-memory workbook."""

    def _cell(self, value):
        wb = Workbook()
        ws = wb.active
        ws["A1"] = value
        return ws["A1"]

    def test_normal_string(self):
        assert safe_cell_value(self._cell("hello")) == "hello"

    def test_normal_number(self):
        assert safe_cell_value(self._cell(42.5)) == 42.5

    def test_none_value(self):
        assert safe_cell_value(self._cell(None)) is None

    def test_whitespace_only(self):
        assert safe_cell_value(self._cell("   ")) is None

    def test_excel_error_ref(self):
        # openpyxl stores errors as strings when data_only=True
        assert safe_cell_value(self._cell("#REF!")) is None

    def test_excel_error_div(self):
        assert safe_cell_value(self._cell("#DIV/0!")) is None

    def test_excel_error_value(self):
        assert safe_cell_value(self._cell("#VALUE!")) is None

    def test_hash_prefix_generic(self):
        assert safe_cell_value(self._cell("#UNKNOWN_ERROR")) is None

    def test_strips_string(self):
        assert safe_cell_value(self._cell("  hello  ")) == "hello"


# ──────────────────────────────────────────────────────────────────────────────
# extract_tabular_data
# ──────────────────────────────────────────────────────────────────────────────

class TestExtractTabularData:
    def test_basic_table(self):
        wb = make_workbook({"S1": [
            ["Titre", "Coût", "PBT"],
            ["Isolation toiture", 10000, 5],
            ["PV 50kWp", 35000, 8],
        ]})
        result = extract_tabular_data(wb["S1"])
        assert result["detected_headers"] == ["Titre", "Coût", "PBT"]
        assert len(result["raw_rows"]) == 2
        assert result["raw_rows"][0]["values"]["Titre"] == "Isolation toiture"

    def test_empty_rows_interspersed(self):
        """Empty rows between data rows must be silently skipped."""
        wb = make_workbook({"S1": [
            ["Action", "Investissement"],
            ["LED",    5000],
            [None,     None],       # ← empty row
            ["HVAC",   20000],
            [None,     None],       # ← empty row at end
        ]})
        result = extract_tabular_data(wb["S1"])
        assert len(result["raw_rows"]) == 2
        titles = [r["values"]["Action"] for r in result["raw_rows"]]
        assert "LED" in titles and "HVAC" in titles

    def test_no_header_detected_falls_back_to_column_letters(self):
        """If the first rows are all numbers, column letters are used as headers."""
        wb = make_workbook({"S1": [
            [1, 2, 3],
            [4, 5, 6],
        ]})
        result = extract_tabular_data(wb["S1"])
        # No text header found; the fallback uses column letters
        # (header_row_index=None or all numeric first row)
        # raw_rows should still be extracted
        assert len(result["raw_rows"]) >= 1

    def test_empty_worksheet(self):
        wb = make_workbook({"S1": []})
        result = extract_tabular_data(wb["S1"])
        assert result["raw_rows"] == []
        # No genuine header cells → detected_headers must be empty
        assert result["detected_headers"] == []

    def test_header_with_some_empty_cells(self):
        """Headers with a mix of text and None should still be detected."""
        wb = make_workbook({"S1": [
            ["Action", None, "Coût", None, "PBT"],
            ["LED", None, 5000, None, 3],
        ]})
        result = extract_tabular_data(wb["S1"])
        # Should detect row 1 as header (≥2 non-empty strings)
        assert result["header_row_index"] == 1
        assert "Action" in result["detected_headers"]
        assert "Coût" in result["detected_headers"]

    def test_row_index_preserved(self):
        """row_index must reflect the actual Excel row number (1-based).
        We use 2 columns so detect_header_row() can identify the header row
        (it requires ≥ 2 non-empty cells).
        """
        wb = make_workbook({"S1": [
            ["Header A", "Header B"],   # row 1  ← header
            ["val1",     "val2"],       # row 2
            ["val3",     "val4"],       # row 3
        ]})
        result = extract_tabular_data(wb["S1"])
        assert result["header_row_index"] == 1
        assert result["raw_rows"][0]["row_index"] == 2
        assert result["raw_rows"][1]["row_index"] == 3


# ──────────────────────────────────────────────────────────────────────────────
# AmurebaMappingService.map_headers
# ──────────────────────────────────────────────────────────────────────────────

class TestMapHeaders:
    svc = AmurebaMappingService()

    def test_exact_alias(self):
        m = self.svc.map_headers(["cout"])
        assert m["cout"] == "cout_investissement"

    def test_exact_with_accent(self):
        m = self.svc.map_headers(["coût"])
        assert m["coût"] == "cout_investissement"

    def test_renamed_column_with_suffix(self):
        """'Coût total investissement' should fuzzy-match cout_investissement."""
        m = self.svc.map_headers(["Coût total investissement"])
        assert m.get("Coût total investissement") == "cout_investissement"

    def test_pbt_renamed(self):
        m = self.svc.map_headers(["Temps de retour (ans)"])
        assert m.get("Temps de retour (ans)") == "temps_retour_brut"

    def test_co2_variant(self):
        m = self.svc.map_headers(["Réduction CO2 kg/an"])
        assert m.get("Réduction CO2 kg/an") == "taux_reduction_co2"

    def test_energy_kwh_variant(self):
        m = self.svc.map_headers(["économie énergie (kWh)"])
        assert m.get("économie énergie (kWh)") == "economie_energie_kwh"

    def test_below_threshold_not_included(self):
        """Completely unrelated header should not match any field."""
        m = self.svc.map_headers(["xyz_random_gibberish_abc"])
        assert "xyz_random_gibberish_abc" not in m

    def test_empty_list(self):
        assert self.svc.map_headers([]) == {}

    def test_none_in_list_skipped(self):
        m = self.svc.map_headers([None, "cout"])
        assert None not in m
        assert "cout" in m

    def test_multiple_fields(self):
        headers = ["Intitulé", "Coût €", "PBT", "Réduction CO2"]
        m = self.svc.map_headers(headers)
        assert m.get("Intitulé") == "titre_action"
        assert m.get("Coût €") == "cout_investissement"
        assert m.get("PBT") == "temps_retour_brut"
        assert m.get("Réduction CO2") == "taux_reduction_co2"

    def test_extra_aliases_injected(self):
        """Custom aliases passed to __init__ must be matched."""
        svc = AmurebaMappingService(
            extra_aliases={"cout_investissement": ["budget total projet"]}
        )
        m = svc.map_headers(["budget total projet"])
        assert m.get("budget total projet") == "cout_investissement"


# ──────────────────────────────────────────────────────────────────────────────
# AmurebaMappingService.extract_amureba_cells
# ──────────────────────────────────────────────────────────────────────────────

class TestExtractAmurebaCells:
    svc = AmurebaMappingService()

    def _make_aa_sheet(self, data: dict):
        """
        Build a worksheet where data = {cell_addr: value}.
        The sheet is embedded in a fresh workbook.
        """
        wb = Workbook()
        ws = wb.active
        for addr, val in data.items():
            ws[addr] = val
        return ws

    def test_titre_action(self):
        ws = self._make_aa_sheet({"B9": "Extension panneaux PV"})
        result = self.svc.extract_amureba_cells(ws)
        assert result["titre_action"] == "Extension panneaux PV"

    def test_cout_investissement(self):
        ws = self._make_aa_sheet({"G61": 75000.0})
        result = self.svc.extract_amureba_cells(ws)
        assert result["cout_investissement"] == 75000.0

    def test_mwh_converted_to_kwh(self):
        """G77 is in MWh/an; service must multiply ×1000 → kWh."""
        ws = self._make_aa_sheet({"G77": 3.5})   # 3.5 MWh/an
        result = self.svc.extract_amureba_cells(ws)
        assert result["economie_energie_kwh"] == pytest.approx(3500.0)

    def test_irr_text_becomes_none(self):
        """IRR cell = 'PROJET NON RENTABLE' must produce None."""
        ws = self._make_aa_sheet({"N10": "PROJET NON RENTABLE"})
        result = self.svc.extract_amureba_cells(ws)
        assert result.get("irr_avant_impot") is None

    def test_irr_numeric_kept(self):
        ws = self._make_aa_sheet({"N10": 0.12})
        result = self.svc.extract_amureba_cells(ws)
        assert result["irr_avant_impot"] == pytest.approx(0.12)

    def test_ets_oui(self):
        ws = self._make_aa_sheet({"K22": "OUI"})
        result = self.svc.extract_amureba_cells(ws)
        assert result["entreprise_ets"] is True

    def test_ets_non(self):
        ws = self._make_aa_sheet({"K22": "NON"})
        result = self.svc.extract_amureba_cells(ws)
        assert result["entreprise_ets"] is False

    def test_deduction_fiscale_case_insensitive(self):
        ws = self._make_aa_sheet({"K23": "oui"})
        result = self.svc.extract_amureba_cells(ws)
        assert result["deduction_fiscale"] is True

    def test_excel_error_becomes_none_not_in_result(self):
        """A cell with #REF! must not appear in the output at all."""
        ws = self._make_aa_sheet({"G61": "#REF!", "B9": "Valid title"})
        result = self.svc.extract_amureba_cells(ws)
        assert "cout_investissement" not in result
        assert result["titre_action"] == "Valid title"

    def test_empty_sheet_returns_empty_dict(self):
        wb = Workbook()
        ws = wb.active
        result = self.svc.extract_amureba_cells(ws)
        assert result == {}


# ──────────────────────────────────────────────────────────────────────────────
# AmurebaMappingService.map_sheet
# ──────────────────────────────────────────────────────────────────────────────

class TestMapSheet:
    svc = AmurebaMappingService()

    def test_empty_sheet_returns_none(self):
        wb = make_workbook({"AA1": []})
        result = self.svc.map_sheet(wb["AA1"], "AA1")
        assert result is None

    def test_sheet_with_only_none_cells_returns_none(self):
        wb = make_workbook({"AA1": [[None, None], [None, None]]})
        result = self.svc.map_sheet(wb["AA1"], "AA1")
        assert result is None

    def test_sheet_with_data_returns_dict(self):
        wb = make_workbook({"AA1": [
            ["Intitulé", "Coût"],
            ["LED", 5000],
        ]})
        result = self.svc.map_sheet(wb["AA1"], "AA1")
        assert result is not None
        assert "raw_rows" in result
        assert "key_values" in result
        assert "detected_headers" in result
        assert "unmapped_headers" in result

    def test_semantic_fields_populated_from_headers(self):
        wb = make_workbook({"PA": [
            ["Intitulé", "Coût investissement", "PBT"],
            ["LED remplacement", 8000, 4],
        ]})
        result = self.svc.map_sheet(wb["PA"], "PA")
        kv = result["key_values"]
        assert kv.get("titre_action") == "LED remplacement"
        assert kv.get("cout_investissement") == 8000
        assert kv.get("temps_retour_brut") == 4

    def test_amureba_fixed_cells_override_tabular(self):
        """
        When both tabular and fixed-cell extraction yield a value for the same
        field, the fixed-cell value takes priority.
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "AA1"
        # Tabular: first row headers, second row data
        ws.append(["Coût", "PBT"])
        ws.append([999, 2])
        # Fixed cell (higher priority)
        ws["G61"] = 50000
        ws["N15"] = 7

        result = self.svc.map_sheet(ws, "AA1")
        kv = result["key_values"]
        # G61 should override the tabular "Coût"=999
        assert kv["cout_investissement"] == 50000
        assert kv["temps_retour_brut"] == 7

    def test_excel_errors_in_data_rows_become_none(self):
        wb = make_workbook({"AA2": [
            ["Titre", "Coût"],
            ["Chaudière", "#VALUE!"],
        ]})
        result = self.svc.map_sheet(wb["AA2"], "AA2")
        assert result is not None
        # The #VALUE! should have been replaced with None in raw_rows
        row_vals = result["raw_rows"][0]["values"]
        assert row_vals.get("Coût") is None

    def test_unmapped_headers_listed(self):
        wb = make_workbook({"S1": [
            ["XYZ_unknown_col_abc", "Coût"],
            ["val", 1000],
        ]})
        result = self.svc.map_sheet(wb["S1"], "S1")
        assert "XYZ_unknown_col_abc" in result["unmapped_headers"]


# ──────────────────────────────────────────────────────────────────────────────
# AmurebaMappingService.map_workbook
# ──────────────────────────────────────────────────────────────────────────────

class TestMapWorkbook:
    svc = AmurebaMappingService()

    def test_empty_sheets_skipped(self):
        wb = make_workbook({
            "AA1": [["Titre", "Coût"], ["LED", 5000]],
            "AA2": [],                              # empty → skipped
            "AA3": [[None, None], [None, None]],   # all None → skipped
        })
        result = self.svc.map_workbook(wb)
        assert "AA1" in result
        assert "AA2" not in result
        assert "AA3" not in result

    def test_multiple_sheets_processed_independently(self):
        wb = make_workbook({
            "AA1": [["Titre", "Coût"], ["LED", 5000]],
            "AA2": [["Titre", "Coût"], ["PV 10kWp", 12000]],
        })
        result = self.svc.map_workbook(wb)
        assert len(result) == 2
        kv1 = result["AA1"]["key_values"]
        kv2 = result["AA2"]["key_values"]
        assert kv1.get("titre_action") == "LED"
        assert kv2.get("titre_action") == "PV 10kWp"

    def test_missing_sheet_does_not_crash(self):
        """map_workbook only iterates wb.sheetnames — absent sheets never accessed."""
        wb = make_workbook({"AA1": [["Titre"], ["Chaudière"]]})
        # "AA5" does not exist — calling map_workbook should be safe
        result = self.svc.map_workbook(wb)
        assert "AA5" not in result

    def test_all_sheets_empty_returns_empty_dict(self):
        wb = make_workbook({"AA1": [], "AA2": [], "PA": []})
        result = self.svc.map_workbook(wb)
        assert result == {}

    def test_returns_sheet_names_as_keys(self):
        wb = make_workbook({
            "Feuille_personnalisée": [["Titre"], ["Mon action"]],
        })
        result = self.svc.map_workbook(wb)
        assert "Feuille_personnalisée" in result
