"""
Tests d'intégration - bibliothèque de modèles de livrable (templates), étape (i).

Base : SQLite en mémoire (cf. conftest). Les FK SQLite sont activées ici pour
vérifier ON DELETE SET NULL (Postgres prod : enforcement natif).
"""
import io
import sqlite3
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from sqlalchemy import event
from sqlalchemy.engine import Engine

from app import models


# ── Active l'enforcement des FK côté SQLite (no-op sur Postgres) ──────────────
@event.listens_for(Engine, "connect")
def _sqlite_fk_pragma(dbapi_connection, connection_record):
    if isinstance(dbapi_connection, sqlite3.Connection):
        cur = dbapi_connection.cursor()
        cur.execute("PRAGMA foreign_keys=ON")
        cur.close()


_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_DOCX_MIME = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"


def _xlsx_with_aa(marker: str | None = None) -> bytes:
    """Workbook avec les 9 feuilles AA1..AA9 (+ feuille marqueur optionnelle)."""
    wb = Workbook()
    wb.remove(wb.active)
    for i in range(1, 10):
        wb.create_sheet(f"AA{i}")
    if marker:
        wb.create_sheet(marker)["A1"] = marker
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xlsx_without_aa() -> bytes:
    wb = Workbook()
    wb.active.title = "Feuil1"
    wb.active["A1"] = "x"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _get_or_create_official(db, ttype: str) -> models.Template:
    tid = f"official-{ttype}"
    existing = db.get(models.Template, tid)
    if existing:
        return existing
    tpl = models.Template(
        id=tid, type=ttype, name=f"Modèle officiel {ttype}",
        file_bytes=None, is_official=True, supports_prefill=True,
        owner_id=None, scope="official", created_at="2026-01-01T00:00:00+00:00",
    )
    db.add(tpl)
    db.commit()
    return tpl


def _upload_audit_custom(client, name="Mon modèle audit", marker=None):
    return client.post(
        "/templates",
        data={"type": "audit", "name": name},
        files={"file": ("m.xlsx", _xlsx_with_aa(marker), _XLSX_MIME)},
    )


# ── Validation à l'upload ────────────────────────────────────────────────────

def test_upload_audit_rejects_missing_aa_sheets(client):
    r = client.post(
        "/templates",
        data={"type": "audit", "name": "Invalide"},
        files={"file": ("x.xlsx", _xlsx_without_aa(), _XLSX_MIME)},
    )
    assert r.status_code == 400
    assert "AA" in r.json()["detail"]


def test_upload_report_rejects_non_docx(client):
    r = client.post(
        "/templates",
        data={"type": "report", "name": "Pas un docx"},
        files={"file": ("notes.txt", b"hello", "text/plain")},
    )
    assert r.status_code == 400


# ── Protection de l'officiel + ownership ─────────────────────────────────────

def test_delete_official_forbidden(client, db_session):
    _get_or_create_official(db_session, "audit")
    r = client.delete("/templates/official-audit")
    assert r.status_code == 403


def test_delete_custom_other_owner_not_found(client, db_session):
    other_id = f"other-user-{uuid4().hex[:8]}"
    db_session.add(models.User(
        id=other_id, full_name="Autrui", email=f"{other_id}@test.internal",
        hashed_password="unused",
    ))
    db_session.commit()
    tpl = models.Template(
        id=f"t-{uuid4().hex[:8]}", type="report", name="Modèle d'autrui",
        file_bytes=b"x", is_official=False, supports_prefill=False,
        owner_id=other_id, scope="user", created_at="2026-01-01T00:00:00+00:00",
    )
    db_session.add(tpl)
    db_session.commit()
    r = client.delete(f"/templates/{tpl.id}")
    assert r.status_code == 404


# ── Sélection du modèle actif (FK) ───────────────────────────────────────────

def test_patch_active_template_sets_fk(client, db_session, seed_project):
    tid = _upload_audit_custom(client).json()["id"]
    r = client.patch(
        f"/projects/{seed_project.id}/active-template",
        json={"type": "audit", "template_id": tid},
    )
    assert r.status_code == 200
    db_session.expire_all()
    proj = db_session.get(models.Project, seed_project.id)
    assert proj.active_audit_template_id == tid


def test_delete_active_custom_resets_project_to_null(client, db_session, seed_project):
    """ON DELETE SET NULL : supprimer le modèle actif repasse le projet à l'officiel (null)."""
    tid = _upload_audit_custom(client).json()["id"]
    client.patch(
        f"/projects/{seed_project.id}/active-template",
        json={"type": "audit", "template_id": tid},
    )
    rd = client.delete(f"/templates/{tid}")
    assert rd.status_code == 200
    db_session.expire_all()
    proj = db_session.get(models.Project, seed_project.id)
    assert proj.active_audit_template_id is None


# ── Génération : export-excel sert le modèle custom actif ────────────────────

def test_export_excel_serves_active_custom(client, db_session, seed_project):
    tid = _upload_audit_custom(client, marker="CUSTOM_MARKER").json()["id"]
    client.patch(
        f"/projects/{seed_project.id}/active-template",
        json={"type": "audit", "template_id": tid},
    )
    r = client.get(f"/projects/{seed_project.id}/improvement-actions/export-excel")
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    assert "CUSTOM_MARKER" in wb.sheetnames   # prouve que le custom (et non l'officiel) a servi


# ── Gating du prefill IA quand un custom est actif (supports_prefill=false) ──

def test_apply_prefill_audit_blocked_for_custom(client, db_session, seed_project):
    tid = _upload_audit_custom(client).json()["id"]
    client.patch(
        f"/projects/{seed_project.id}/active-template",
        json={"type": "audit", "template_id": tid},
    )
    r = client.post(
        f"/projects/{seed_project.id}/improvement-actions/apply-prefill",
        json={"changes": []},
    )
    assert r.status_code == 409


def test_apply_prefill_report_blocked_for_custom(client, db_session, seed_project):
    tid = client.post(
        "/templates",
        data={"type": "report", "name": "Mon rapport"},
        files={"file": ("r.docx", b"PK-fake-docx-bytes", _DOCX_MIME)},
    ).json()["id"]
    client.patch(
        f"/projects/{seed_project.id}/active-template",
        json={"type": "report", "template_id": tid},
    )
    r = client.post(
        f"/projects/{seed_project.id}/report/apply-prefill",
        json={"items": []},
    )
    assert r.status_code == 409
